import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from datetime import datetime
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Count, Sum, Q
from datetime import datetime, date, timedelta
from farmacia.models import Venta, DetalleVenta, Medicamento
from django.http import JsonResponse
from datetime import datetime
from .models import Consulta
from config.decorators import grupo_requerido
from consultas.models import CatalogoExamen, CatalogoEcografia, OrdenExamen
from .models import Consulta, CatalogoExamen, CatalogoEcografia, TopicoCatalogo, OrdenExamen, VentaAmbulatoria
from datetime import timedelta
from decimal import Decimal
from django.views.decorators.http import require_POST
from django.shortcuts import render
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.shortcuts import render, get_object_or_404, redirect
from pacientes.models import Paciente
from .models import Consulta, Receta, OrdenExamen
from .forms import ConsultaForm
from django.utils.timezone import now
from pagos.models import OrdenPago
from doctores.models import Doctor
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils import timezone
from django.contrib import messages
from .models import CatalogoExamen
from django.db.models import Q
from .models import (
    Consulta,
    Receta,
    OrdenExamen,
    CatalogoExamen,
    VentaAmbulatoria
)
from notificaciones.models import Notificacion
from django.contrib.auth.models import User

##########################
# FUNCIONES AUXILIARES
##########################
@login_required
@grupo_requerido('doctores', 'laboratorio', 'administrador', 'recepcion')
def calcular_estadisticas_ordenes(request):
    """
    Calcula estadísticas CORRECTAS para órdenes de laboratorio.
    SIN contar urgentes (según requerimiento).
    """
    # Obtener todas las órdenes de laboratorio pendientes
    ordenes_pendientes = OrdenExamen.objects.filter(
        tipo_examen='LABORATORIO',
        estado='SOLICITADO'
    )

    # Calcular contadores REALES (sin urgentes)
    stats = {
        'total_pendientes': ordenes_pendientes.count(),
        'internos': 0,
        'ambulatorios': 0,
    }

    # Calcular usando Python
    for orden in ordenes_pendientes:
        if orden.consulta:
            stats['internos'] += 1
        else:
            stats['ambulatorios'] += 1

    return stats
@login_required
@grupo_requerido('doctores','laboratorio','administrador')
def get_contadores_ordenes(request):
    """Obtiene contadores para todos los estados"""
    return {
        'pendientes_count': OrdenExamen.objects.filter(
            tipo_examen='LABORATORIO',
            estado='SOLICITADO'
        ).count(),
        'en_proceso_count': OrdenExamen.objects.filter(
            tipo_examen='LABORATORIO',
            estado='EN_PROCESO'
        ).count(),
        'completadas_count': OrdenExamen.objects.filter(
            tipo_examen='LABORATORIO',
            estado='COMPLETADO'
        ).count(),
        'entregadas_count': OrdenExamen.objects.filter(
            tipo_examen='LABORATORIO',
            estado='ENTREGADO'
        ).count(),
    }
@login_required
@grupo_requerido('laboratorio')
def es_laboratorio(user):
    """Verifica si el usuario pertenece al grupo laboratorio"""
    return user.groups.filter(name='laboratorio').exists()

@login_required
@grupo_requerido('doctores','laboratorio','administrador')
def obtener_mis_ordenes_simplificado(user):
    """
    Obtiene mis órdenes en proceso SIN información de urgente/prioridad/tiempo
    según requerimiento 5.
    """
    # Solo órdenes asignadas al usuario actual
    ordenes = OrdenExamen.objects.filter(
        tipo_examen='LABORATORIO',
        estado='EN_PROCESO',
        tecnico_asignado=user
    ).select_related(
        'consulta__paciente',
        'consulta__doctor__usuario',
        'venta_ambulatoria'
    ).order_by('fecha_asignacion')

    # NO contamos urgentes, NO calculamos tiempos
    stats = {
        'total': ordenes.count(),
        'internos': ordenes.filter(consulta__isnull=False).count(),
        'ambulatorios': ordenes.filter(consulta__isnull=True).count(),
    }

    return ordenes, stats

# ==============================================
# VISTAS PARA LABORATORIO
# ==============================================
@login_required
@grupo_requerido('doctores', 'laboratorio', 'administrador', 'recepcion')
def lista_ordenes_pendientes(request):
    """
    Vista para órdenes pendientes - CON ESTADO DE PAGO CORREGIDO
    """
    # 1. Órdenes pendientes
    ordenes_pendientes = OrdenExamen.objects.filter(
        tipo_examen='LABORATORIO',
        estado='SOLICITADO'
    ).select_related(
        'consulta__paciente',
        'venta_ambulatoria'
    ).order_by('fecha_solicitud')

    # ✅ CORRECCIÓN: Determinar estado de pago para cada orden CON MANEJO DE None
    for orden in ordenes_pendientes:
        # ✅ SOLUCIÓN CRÍTICA: Asegurar que pagado no sea None
        if orden.pagado is None:
            orden.pagado = False
        
        if orden.consulta is None:  # Es ambulatorio
            orden.estado_pago_display = '✅ Pagado'
            orden.puede_cobrar = False
            orden.es_pagado = True
        else:  # Es interno
            if orden.pagado:  # Ya pagó
                orden.estado_pago_display = '✅ Pagado'
                orden.puede_cobrar = False
                orden.es_pagado = True
            else:  # No ha pagado
                orden.estado_pago_display = '❌ Por Cobrar'
                orden.puede_cobrar = True
                orden.es_pagado = False

    # 2. Órdenes para las otras pestañas
    ordenes_en_proceso = OrdenExamen.objects.filter(
        tipo_examen='LABORATORIO',
        estado='EN_PROCESO'
    ).select_related(
        'consulta__paciente',
        'venta_ambulatoria'
    ).order_by('-fecha_asignacion')

    ordenes_completados = OrdenExamen.objects.filter(
        tipo_examen='LABORATORIO',
        estado='COMPLETADO'
    ).select_related(
        'consulta__paciente',
        'venta_ambulatoria'
    ).order_by('-fecha_realizacion')

    ordenes_entregados = OrdenExamen.objects.filter(
        tipo_examen='LABORATORIO',
        estado='ENTREGADO'
    ).select_related(
        'consulta__paciente',
        'venta_ambulatoria'
    ).order_by('-fecha_entrega')

    # 3. Calcular estadísticas
    stats = calcular_estadisticas_ordenes(request)
    contadores = get_contadores_ordenes(request)
    # 4. Contexto COMPLETO
    context = {
        # Variables principales
        'ordenes_pendientes': ordenes_pendientes,
        'ordenes_en_proceso': ordenes_en_proceso,
        'ordenes_completados': ordenes_completados,
        'ordenes_entregados': ordenes_entregados,

        'titulo': 'Pendientes',
        'es_laboratorio': True,

        # Estadísticas
        'internos_count': stats['internos'],
        'ambulatorios_count': stats['ambulatorios'],
        'total_pendientes': stats['total_pendientes'],

        # Contadores
        'pendientes_count': contadores.get('pendientes_count', 0),
        'en_proceso_count': contadores.get('en_proceso_count', 0),
        'completadas_count': contadores.get('completadas_count', 0),
        'entregadas_count': contadores.get('entregadas_count', 0),

        # Para búsqueda de entregados (si aplica)
        'dni_busqueda': request.GET.get('dni', ''),
    }

    return render(request, 'consultas/laboratorio/lista_ordenes.html', context)

######################################
@login_required
@grupo_requerido('laboratorio', 'administrador')
def mis_ordenes_en_proceso(request):
    """
    Muestra solo las órdenes asignadas AL USUARIO ACTUAL
    """
    # Órdenes de consulta (internos) en proceso
    ordenes_internos = OrdenExamen.objects.filter(
        tipo_examen='LABORATORIO',
        estado='EN_PROCESO',
        tecnico_asignado=request.user,
        consulta__isnull=False
    ).select_related(
        'consulta__paciente',
        'consulta__doctor__usuario'
    ).order_by('fecha_solicitud')

    # Órdenes ambulatorias en proceso
    ordenes_ambulatorios = OrdenExamen.objects.filter(
        tipo_examen='LABORATORIO',
        estado='EN_PROCESO',
        tecnico_asignado=request.user,
        consulta__isnull=True,
        venta_ambulatoria__isnull=False
    ).select_related('venta_ambulatoria').order_by('fecha_solicitud')

    # Combinar ambas listas
    ordenes = list(ordenes_internos) + list(ordenes_ambulatorios)

    # Contar órdenes urgentes
    ordenes_urgentes_count = 0 # sum(1 for orden in ordenes if orden.urgente)

    context = {
        'ordenes': ordenes,
        'ordenes_internos_count': len(ordenes_internos),
        'ordenes_ambulatorios_count': len(ordenes_ambulatorios),
        'ordenes_urgentes_count': ordenes_urgentes_count,
        'es_laboratorio': True,
        **get_contadores_ordenes(request),
    }

    return render(request, 'consultas/laboratorio/mis_ordenes.html', context)

@login_required
@grupo_requerido('doctores', 'laboratorio', 'administrador', 'recepcion')
def ordenes_en_proceso(request):
    """Órdenes en estado EN_PROCESO (para todos los técnicos)"""
    ordenes = OrdenExamen.objects.filter(
        tipo_examen='LABORATORIO',
        estado='EN_PROCESO'
    ).select_related(
        'consulta__paciente',
        'venta_ambulatoria',
        'tecnico_asignado'
    ).order_by('-fecha_asignacion')

    contadores = get_contadores_ordenes(request)
    
    context = {
        'ordenes_en_proceso': ordenes,  # ¡IMPORTANTE! Nombre correcto
        'titulo': 'En Proceso',
        'es_laboratorio': True,
        **contadores,
    }

    return render(request, 'consultas/laboratorio/lista_ordenes.html', context)

@login_required
@grupo_requerido('doctores', 'laboratorio', 'administrador', 'recepcion')
def ordenes_completadas(request):
    """Órdenes en estado COMPLETADO"""
    ordenes = OrdenExamen.objects.filter(
        tipo_examen='LABORATORIO',
        estado='COMPLETADO'
    ).select_related(
        'consulta__paciente',
        'venta_ambulatoria',
        'tecnico_asignado'
    ).order_by('-fecha_realizacion')

    contadores = get_contadores_ordenes(request)
    
    # Calcular estadísticas adicionales
    total_completados = ordenes.count()
    internos_completados = ordenes.filter(consulta__isnull=False).count()
    ambulatorios_completados = ordenes.filter(consulta__isnull=True).count()
    
    # Completados hoy
    hoy = timezone.now().date()
    completados_hoy = ordenes.filter(fecha_realizacion__date=hoy).count()

    context = {
        'ordenes_completados': ordenes,  # ¡IMPORTANTE! Nombre correcto (PLURAL)
        'titulo': 'Completadas',
        'es_laboratorio': True,
        'total_completados': total_completados,
        'internos_completados': internos_completados,
        'ambulatorios_completados': ambulatorios_completados,
        'completados_hoy': completados_hoy,
        **contadores,
    }

    return render(request, 'consultas/laboratorio/lista_ordenes.html', context)
###############################
# En views.py, reemplaza la función ordenes_entregadas con:
@login_required
@grupo_requerido('doctores', 'laboratorio', 'administrador', 'recepcion')
def ordenes_entregadas(request):
    """Órdenes en estado ENTREGADO - VERSIÓN SIMPLE"""
    # Solo obtener órdenes entregadas
    ordenes = OrdenExamen.objects.filter(
        tipo_examen='LABORATORIO',
        estado='ENTREGADO'
    ).select_related(
        'consulta__paciente',
        'venta_ambulatoria'
    ).order_by('-fecha_entrega')
    
    # Paginación simple
    from django.core.paginator import Paginator
    paginator = Paginator(ordenes, 15)  # 15 por página
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Contadores simples (solo para badges de tabs)
    contadores = get_contadores_ordenes(request)
    
    context = {
        'ordenes_entregados': page_obj,
        'page_obj': page_obj,
        'titulo': 'Entregadas',
        
        # Solo contadores necesarios para badges
        'pendientes_count': contadores.get('pendientes_count', 0),
        'en_proceso_count': contadores.get('en_proceso_count', 0),
        'completadas_count': contadores.get('completadas_count', 0),
        'entregadas_count': contadores.get('entregadas_count', 0),
    }
    
    return render(request, 'consultas/laboratorio/lista_ordenes.html', context)

##################
@login_required
@grupo_requerido('laboratorio', 'administrador')
def asignar_orden(request, orden_id):
    """Asignar una orden específica al laboratorista actual"""
    print(f"\n{'='*50}")
    print(f"DEBUG ASIGNAR_ORDEN - Orden #{orden_id}")
    
    orden = get_object_or_404(OrdenExamen, id=orden_id)
    
    # DEBUG
    print(f"Estado: {orden.estado}")
    print(f"Técnico asignado: {orden.tecnico_asignado}")
    print(f"Pagado: {orden.pagado}")
    print(f"Es interno: {orden.consulta is not None}")
    
    # VALIDACIONES MEJORADAS
    if orden.tipo_examen != 'LABORATORIO':
        print("❌ No es laboratorio")
        messages.error(request, '❌ Solo se pueden asignar órdenes de Laboratorio.')
        return redirect('lista_ordenes_pendientes')

    if orden.estado != 'SOLICITADO':
        print(f"❌ Estado incorrecto: {orden.estado}")
        estado_actual = orden.get_estado_display()
        messages.error(request, f'❌ La orden ya está {estado_actual.lower()}.')
        return redirect('lista_ordenes_pendientes')

    if orden.tecnico_asignado:
        print(f"❌ Ya tiene técnico: {orden.tecnico_asignado}")
        tecnico = orden.tecnico_asignado.get_full_name() or orden.tecnico_asignado.username
        messages.warning(request, f'⚠️ Ya está asignada a: {tecnico}')
        return redirect('lista_ordenes_pendientes')

    # ✅ NUEVA VALIDACIÓN: Si es INTERNA, debe estar PAGADA
    if orden.consulta and not orden.pagado:
        print(f"❌ Orden interna no pagada")
        messages.error(request, '❌ Primero debe cobrar la orden interna antes de asignarla.')
        return redirect('lista_ordenes_pendientes')

    # ASIGNAR
    print("✅ Validaciones OK - Asignando...")
    orden.estado = 'EN_PROCESO'
    orden.tecnico_asignado = request.user
    orden.fecha_asignacion = timezone.now()
    orden.save()
    
    print(f"✅ Orden #{orden.id} asignada a {request.user.username}")
    print(f"{'='*50}\n")
    
    user_name = request.user.get_full_name() or request.user.username
    messages.success(request, f'✅ Orden #{orden.id} asignada a: {user_name}')

    return redirect('mis_ordenes_en_proceso')
##################
@login_required
@grupo_requerido('doctores', 'laboratorio', 'administrador', 'recepcion')
def detalle_orden_examen(request, orden_id):
    """
    Vista CORREGIDA para ver detalles de una orden de examen
    """
    orden = get_object_or_404(
        OrdenExamen.objects.select_related(
            'consulta__paciente',
            'consulta__doctor',
            'venta_ambulatoria'
        ),
        id=orden_id,
        tipo_examen='LABORATORIO'
    )

    es_ambulatoria = orden.consulta is None and orden.venta_ambulatoria is not None

    # ✅ NUEVO: Verificar si la orden se puede asignar
    puede_asignarse = (
        orden.estado == 'SOLICITADO' and 
        not orden.pagado and  # ← ¡IMPORTANTE! Verificar que NO esté pagada
        not orden.tecnico_asignado
    )
    
    # ✅ NUEVO: Motivo por el que no se puede asignar
    motivo_no_asignable = ""
    if orden.pagado:
        motivo_no_asignable = "Por pagar - Complete el pago primero"
    elif orden.tecnico_asignado:
        motivo_no_asignable = f"Ya asignada a {orden.tecnico_asignado.get_full_name()}"
    elif orden.estado != 'SOLICITADO':
        motivo_no_asignable = f"Estado actual: {orden.get_estado_display()}"

    # PREPARAR DATOS DEL PACIENTE (código existente sin cambios)
    paciente_data = {}
    if es_ambulatoria:
        venta = orden.venta_ambulatoria
        paciente_data = {
            'nombres': venta.paciente_nombre if venta else 'N/A',
            'apellidos': '',
            'nombre_completo': venta.paciente_nombre if venta else 'N/A',
            'dni': venta.paciente_dni if venta else 'N/A',
            'edad': 'No disponible',
            'telefono': 'No disponible',
            'es_ambulatorio': True,
        }
    else:
        if orden.consulta and orden.consulta.paciente:
            paciente = orden.consulta.paciente

            edad_info = 'N/A'
            try:
                if hasattr(paciente, 'get_edad'):
                    edad_info = paciente.get_edad()
                elif hasattr(paciente, 'edad'):
                    edad_val = paciente.edad
                    if callable(edad_val):
                        edad_info = edad_val()
                    else:
                        edad_info = str(edad_val)
            except:
                edad_info = 'N/A'

            telefono_info = 'No registrado'
            try:
                if hasattr(paciente, 'telefono'):
                    telefono = paciente.telefono
                    if telefono:
                        telefono_info = telefono
            except:
                telefono_info = 'No registrado'

            paciente_data = {
                'nombres': paciente.nombres if hasattr(paciente, 'nombres') else 'N/A',
                'apellidos': paciente.apellidos if hasattr(paciente, 'apellidos') else '',
                'nombre_completo': f"{paciente.nombres} {paciente.apellidos}".strip() if hasattr(paciente, 'nombres') else 'N/A',
                'dni': paciente.dni if hasattr(paciente, 'dni') else 'N/A',
                'edad': edad_info,
                'telefono': telefono_info,
                'es_ambulatorio': False,
            }
        else:
            paciente_data = {
                'nombres': 'N/A',
                'apellidos': '',
                'nombre_completo': 'Datos no disponibles',
                'dni': 'N/A',
                'edad': 'N/A',
                'telefono': 'No disponible',
                'es_ambulatorio': False,
            }

    # MANEJO DE PETICIONES POST (código existente sin cambios)
    if request.method == 'POST':
        accion = request.POST.get('accion')

        if accion == 'completar':
            orden.estado = 'COMPLETADO'
            orden.fecha_realizacion = timezone.now()
            orden.save()
            messages.success(request, f'✅ Orden #{orden.id} marcada como completada')
            return redirect('ordenes_completadas')

        elif accion == 'entregar':
            orden.estado = 'ENTREGADO'
            orden.fecha_entrega = timezone.now()
            orden.save()
            messages.success(request, f'✅ Orden #{orden.id} marcada como entregada')
            return redirect('ordenes_entregadas')

        return redirect('detalle_orden_examen', orden_id=orden.id)

    context = {
        'orden': orden,
        'paciente': paciente_data,
        'consulta': orden.consulta,
        'es_ambulatoria': es_ambulatoria,
        'es_laboratorio': True,
        'puede_asignarse': puede_asignarse,  # ← NUEVO
        'motivo_no_asignable': motivo_no_asignable,  # ← NUEVO
    }

    return render(request, 'consultas/laboratorio/detalle_orden.html', context)

##################
@login_required
@grupo_requerido('laboratorio')
def cancelar_orden(request, orden_id):
    """
    Cancelar una orden en estado SOLICITADO
    """
    orden = get_object_or_404(OrdenExamen, id=orden_id)

    if orden.estado != 'SOLICITADO':
        messages.error(request, f'❌ Solo se pueden cancelar órdenes en estado "Solicitado". Esta orden ya está {orden.get_estado_display()}.')
        return redirect('lista_ordenes_pendientes')

    orden_id = orden.id
    paciente_nombre = ""

    if orden.consulta:
        paciente_nombre = f"{orden.consulta.paciente.nombres} {orden.consulta.paciente.apellidos}"
    elif orden.venta_ambulatoria:
        paciente_nombre = orden.venta_ambulatoria.paciente_nombre

    orden.delete()

    messages.success(request, f'✅ Orden #{orden_id} cancelada exitosamente. Paciente: {paciente_nombre}')
    return redirect('lista_ordenes_pendientes')

# ==============================================
# VISTAS PARA PUNTO DE VENTA LABORATORIO
# ==============================================
@login_required
@grupo_requerido('laboratorio', 'administrador')
def punto_venta_laboratorio(request):
    examenes = CatalogoExamen.objects.filter(
        activo=True,
        tipo='LABORATORIO'
    ).order_by('nombre')

    carrito = request.session.get('carrito_laboratorio', {})
    
    # NUEVO: Calcular tipos en el carrito
    tiene_examenes_nuevos = False
    tiene_ordenes_internas = False
    
    for item in carrito.values():
        if item.get('tipo') == 'INTERNO_POR_COBRAR':
            tiene_ordenes_internas = True
        else:
            tiene_examenes_nuevos = True
    
    total = Decimal('0.00')
    for item in carrito.values():
        total += Decimal(item['precio']) * item['cantidad']

    context = {
        'examenes': examenes,
        'carrito': carrito,
        'total_carrito': total,
        'items_carrito': len(carrito),
        # NUEVAS VARIABLES:
        'tiene_examenes_nuevos': tiene_examenes_nuevos,
        'tiene_ordenes_internas': tiene_ordenes_internas,
        'es_laboratorio': True,
    }

    return render(request, 'consultas/laboratorio/punto_venta.html', context)
################################################
@login_required
@grupo_requerido('laboratorio', 'administrador')
def agregar_al_carrito(request, examen_id):
    """Agregar un examen al carrito - CON VALIDACIÓN DE NO MEZCLA"""
    
    # ✅ VALIDACIÓN SIMPLE: Si tiene órdenes internas, NO agregar examen nuevo
    carrito = request.session.get('carrito_laboratorio', {})
    tiene_ordenes_internas = any(k.startswith('orden_interna_') for k in carrito.keys())
    
    if tiene_ordenes_internas:
        messages.error(request, '❌ No puedes agregar exámenes nuevos porque ya tienes órdenes internas en el carrito.')
        return redirect('punto_venta_laboratorio')
    
    # RESTO DE TU CÓDIGO ORIGINAL (sin cambios)
    examen = get_object_or_404(CatalogoExamen, id=examen_id, activo=True)

    if 'carrito_laboratorio' not in request.session:
        request.session['carrito_laboratorio'] = {}

    carrito = request.session['carrito_laboratorio']

    if str(examen_id) in carrito:
        carrito[str(examen_id)]['cantidad'] += 1
    else:
        carrito[str(examen_id)] = {
            'id': examen.id,
            'nombre': examen.nombre,
            'tipo': examen.tipo,
            'precio': str(examen.precio),
            'cantidad': 1,
            'descripcion': examen.descripcion,
        }

    request.session.modified = True
    messages.success(request, f"✅ {examen.nombre} agregado al carrito")
    return redirect('punto_venta_laboratorio')

###################################
# Asegúrate que recibe 'item_id' (no 'examen_id')
@login_required
@grupo_requerido('laboratorio', 'administrador')
def eliminar_del_carrito(request, item_id):  # <-- item_id, no examen_id
    """Elimina un item del carrito"""
    carrito = request.session.get('carrito_laboratorio', {})
    
    if str(item_id) in carrito:  # <-- Convertir a string por seguridad
        del carrito[str(item_id)]
        request.session['carrito_laboratorio'] = carrito
        messages.success(request, 'Item eliminado del carrito')
    
    return redirect('punto_venta_laboratorio')
##########################################################
@login_required
@grupo_requerido('laboratorio')
def actualizar_cantidad(request, item_id):  # <-- CAMBIAR examen_id → item_id
    """Actualizar cantidad de un item en el carrito (examen nuevo O orden interna)"""
    if request.method == 'POST':
        cantidad = int(request.POST.get('cantidad', 1))
        
        if cantidad < 1:
            cantidad = 1
        
        carrito = request.session.get('carrito_laboratorio', {})
        
        if str(item_id) in carrito:  # <-- item_id ya es string
            carrito[str(item_id)]['cantidad'] = cantidad
            request.session['carrito_laboratorio'] = carrito
            messages.success(request, 'Cantidad actualizada')
    
    return redirect('punto_venta_laboratorio')

###########
@login_required
@grupo_requerido('laboratorio', 'administrador')
def limpiar_carrito(request):
    """Limpiar todo el carrito"""
    if 'carrito_laboratorio' in request.session:
        del request.session['carrito_laboratorio']
        request.session.modified = True
        messages.info(request, "🗑️ Carrito limpiado")

    return redirect('punto_venta_laboratorio')
#######################################
@login_required
@grupo_requerido('laboratorio', 'administrador')
def procesar_venta_ambulatoria(request):
    """Procesa venta ambulatoria Y cobro de órdenes internas - VERSIÓN CON DESCUENTO PARA AMBOS"""
    if request.method == 'POST':
        dni = request.POST.get('dni', '').strip()
        nombre = request.POST.get('nombre', '').strip()
        observaciones = request.POST.get('observaciones', '').strip()
        metodo_pago = request.POST.get('metodo_pago', 'EFECTIVO')

        # OBTENER DESCUENTO
        descuento_str = request.POST.get('descuento', '0').strip()
        descuento = Decimal(descuento_str) if descuento_str else Decimal('0')

        # Validar descuento
        if descuento < 0:
            descuento = Decimal('0')
            messages.warning(request, '⚠️ El descuento no puede ser negativo. Se estableció en 0.')

        carrito = request.session.get('carrito_laboratorio', {})

        if not carrito:
            messages.error(request, '❌ El carrito está vacío')
            return redirect('punto_venta_laboratorio')

        # SEPARAR ITEMS POR TIPO
        items_nuevos = [item for item in carrito.values() if item.get('tipo') != 'INTERNO_POR_COBRAR']
        items_por_cobrar = [item for item in carrito.values() if item.get('tipo') == 'INTERNO_POR_COBRAR']

        # VARIABLES PARA CALCULAR TOTALES
        total_nuevos = Decimal('0.00')
        total_internas = Decimal('0.00')
        total_sin_descuento = Decimal('0.00')  # Total antes de descuento
        ordenes_creadas = []
        ordenes_cobradas = []
        venta = None

        try:
            # 1. CALCULAR TOTALES ANTES DE PROCESAR
            # Calcular total para exámenes nuevos
            for item in items_nuevos:
                total_nuevos += Decimal(str(item['precio'])) * item['cantidad']
            
            # Calcular total para órdenes internas
            for item in items_por_cobrar:
                total_internas += Decimal(str(item['precio']))  # Solo 1 por orden interna
            
            total_sin_descuento = total_nuevos + total_internas
            
            # 2. VALIDAR Y APLICAR DESCUENTO
            porcentaje_descuento = Decimal('0')
            descuento_nuevos = Decimal('0')
            descuento_internas = Decimal('0')
            
            if descuento > 0:
                # No permitir descuento mayor al total
                if descuento > total_sin_descuento:
                    descuento = total_sin_descuento
                    messages.warning(request, f'⚠️ El descuento se ajustó a S/ {descuento:.2f} (no puede superar el total)')
                
                # Calcular porcentaje de descuento
                porcentaje_descuento = descuento / total_sin_descuento if total_sin_descuento > 0 else Decimal('0')
                
                # Aplicar descuento proporcionalmente
                descuento_nuevos = total_nuevos * porcentaje_descuento
                descuento_internas = total_internas * porcentaje_descuento
                
                total_nuevos -= descuento_nuevos
                total_internas -= descuento_internas
                
                print(f"[DEBUG PROCESAR] Descuento total: S/ {descuento:.2f}")
                print(f"[DEBUG PROCESAR] Descuento nuevos: S/ {descuento_nuevos:.2f}")
                print(f"[DEBUG PROCESAR] Descuento internas: S/ {descuento_internas:.2f}")
                print(f"[DEBUG PROCESAR] Total con descuento: S/ {(total_nuevos + total_internas):.2f}")

            # 3. PROCESAR EXÁMENES NUEVOS (AMBULATORIOS)
            if items_nuevos:
                if not dni or not nombre:
                    messages.error(request, '❌ Para exámenes nuevos debe ingresar DNI y nombre del paciente')
                    return redirect('punto_venta_laboratorio')

                # Crear venta primero
                venta = VentaAmbulatoria.objects.create(
                    paciente_dni=dni,
                    paciente_nombre=nombre,
                    total=float(total_nuevos + total_internas),  # Total CON descuento aplicado
                    descuento_aplicado=float(descuento),  # Descuento total aplicado
                    usuario=request.user
                )

                # Crear órdenes para exámenes nuevos
                for item in items_nuevos:
                    examen = CatalogoExamen.objects.get(id=item['id'])
                    precio_examen_original = Decimal(str(item['precio']))
                    
                    # Aplicar descuento proporcional a cada examen nuevo
                    factor_descuento = (Decimal('1') - porcentaje_descuento) if descuento > 0 else Decimal('1')
                    precio_examen_con_descuento = precio_examen_original * factor_descuento

                    for i in range(item['cantidad']):
                        orden_examen = OrdenExamen.objects.create(
                            consulta=None,
                            venta_ambulatoria=venta,
                            tipo_examen='LABORATORIO',
                            examen_especifico=examen.nombre,
                            indicaciones=(
                                f"PACIENTE EXTERNO (AMBULATORIO)\n"
                                f"Nombre: {nombre}\n"
                                f"DNI: {dni}\n"
                                f"Venta: #{venta.id}\n"
                                f"{'Observaciones: ' + observaciones if observaciones else ''}"
                            ),
                            estado='SOLICITADO',
                            pagado=True,
                            metodo_pago=metodo_pago,
                            monto_pagado=float(precio_examen_con_descuento),  # Precio con descuento
                            fecha_pago=timezone.now()
                        )
                        ordenes_creadas.append(orden_examen.id)

            # 4. PROCESAR ÓRDENES INTERNAS POR COBRAR
            if items_por_cobrar:
                for item in items_por_cobrar:
                    try:
                        orden = OrdenExamen.objects.get(id=item['orden_id'])
                        precio_original = Decimal(str(item['precio']))
                        
                        # Aplicar descuento proporcional a orden interna
                        if descuento > 0:
                            factor_descuento = (Decimal('1') - porcentaje_descuento)
                            precio_con_descuento = precio_original * factor_descuento
                        else:
                            precio_con_descuento = precio_original

                        # Marcar como pagada con precio con descuento
                        orden.pagado = True
                        orden.metodo_pago = metodo_pago
                        orden.monto_pagado = float(precio_con_descuento)  # Precio con descuento
                        orden.fecha_pago = timezone.now()
                        orden.save()

                        ordenes_cobradas.append({
                            'id': orden.id,
                            'examen': orden.examen_especifico,
                            'precio_original': float(precio_original),  # Precio original para el ticket
                            'precio': float(precio_con_descuento),  # Precio con descuento (para compatibilidad)
                            'precio_final': float(precio_con_descuento),  # Precio final con descuento
                            'descuento_aplicado': float(precio_original - precio_con_descuento),
                            'paciente': orden.get_paciente_nombre(),
                            'dni': orden.get_paciente_dni()
                        })

                    except OrdenExamen.DoesNotExist:
                        continue

                # Si solo hay órdenes internas (sin exámenes nuevos), crear venta para el ticket
                if not venta and ordenes_cobradas:
                    primer_orden = ordenes_cobradas[0]
                    
                    venta = VentaAmbulatoria.objects.create(
                        paciente_dni=primer_orden['dni'],
                        paciente_nombre=f"COBRO INTERNO - {primer_orden['paciente'][:30]}",
                        total=float(total_internas),  # Total de órdenes internas con descuento
                        descuento_aplicado=float(descuento_internas),
                        usuario=request.user
                    )

            # 5. LIMPIAR CARRITO
            if 'carrito_laboratorio' in request.session:
                del request.session['carrito_laboratorio']
            request.session.modified = True

            # 6. Preparar datos para sesión (IMPORTANTE: para ticket_cobro_interno)
            datos_cobro = {}
            if ordenes_cobradas:
                # Calcular el descuento que realmente se aplicó a las órdenes internas
                descuento_aplicado_internas = descuento_internas if items_por_cobrar else Decimal('0')
                
                datos_cobro = {
                    'ids': [oc['id'] for oc in ordenes_cobradas],
                    'detalles': ordenes_cobradas,
                    'total': float(total_sin_descuento),  # Total sin descuento
                    'descuento': float(descuento_aplicado_internas),  # Descuento aplicado a internas
                    'total_con_descuento': float(total_internas),  # Total con descuento (ya calculado)
                    'metodo_pago': metodo_pago,
                    'fecha': timezone.now().isoformat()
                }
                request.session['ultimas_ordenes_cobradas'] = datos_cobro
                request.session.modified = True
                
                print(f"[DEBUG PROCESAR] Descuento guardado en sesión: {descuento_aplicado_internas}")
                print(f"[DEBUG PROCESAR] Total guardado en sesión: {total_sin_descuento}")
                print(f"[DEBUG PROCESAR] Total con descuento guardado: {total_internas}")

            # 7. MENSAJE COMBINADO
            mensaje_parts = []

            if ordenes_creadas:
                mensaje_parts.append(f'📋 {len(ordenes_creadas)} orden(es) nueva(s) creada(s)')

            if ordenes_cobradas:
                mensaje_parts.append(f'💰 {len(ordenes_cobradas)} orden(es) interna(s) cobrada(s)')

            mensaje = '✅ ' + ' | '.join(mensaje_parts)
            mensaje += f'\n💵 Subtotal: S/ {total_sin_descuento:.2f}'

            if descuento > 0:
                mensaje += f'\n🎫 Descuento: S/ {descuento:.2f}'
                mensaje += f'\n💳 Total con descuento: S/ {(total_sin_descuento - descuento):.2f}'
            else:
                mensaje += f'\n💳 Total: S/ {total_sin_descuento:.2f}'

            messages.success(request, mensaje)

            # 8. REDIRIGIR SEGÚN QUÉ SE PROCESÓ
            if venta:
                if ordenes_creadas:
                    # Hay exámenes nuevos ambulatorios
                    return redirect('ticket_venta_ambulatoria', venta_id=venta.id)
                elif ordenes_cobradas:
                    # Solo órdenes internas cobradas
                    return redirect('ticket_cobro_interno', venta_id=venta.id)
            else:
                return redirect('lista_ordenes_pendientes')

        except CatalogoExamen.DoesNotExist as e:
            messages.error(request, f'❌ Error: Examen no encontrado en el catálogo')
            return redirect('punto_venta_laboratorio')
        except Exception as e:
            messages.error(request, f'❌ Error al procesar: {str(e)}')
            import traceback
            traceback.print_exc()
            return redirect('punto_venta_laboratorio')

    return redirect('punto_venta_laboratorio')
#######################################
@login_required
@grupo_requerido('laboratorio', 'administrador')
def ticket_venta_ambulatoria(request, venta_id):
    """Mostrar ticket térmico para venta ambulatoria CON PRECIOS"""
    try:
        venta = VentaAmbulatoria.objects.get(id=venta_id)
        ordenes_examen = OrdenExamen.objects.filter(venta_ambulatoria=venta)
        
        # NUEVO: Crear lista con exámenes y sus precios
        examenes_con_precio = []
        total_verificado = 0
        
        for orden in ordenes_examen:
            # Buscar el precio del examen en el catálogo
            try:
                examen_catalogo = CatalogoExamen.objects.get(
                    nombre=orden.examen_especifico,
                    tipo='LABORATORIO',
                    activo=True
                )
                precio = examen_catalogo.precio
            except CatalogoExamen.DoesNotExist:
                # Si no encuentra en catálogo, usar monto_pagado o precio por defecto
                precio = orden.monto_pagado if hasattr(orden, 'monto_pagado') else Decimal('50.00')
            
            examenes_con_precio.append({
                'nombre': orden.examen_especifico,
                'precio_unitario': precio,
                'cantidad': 1,  # Cada orden es un examen
                'subtotal': precio
            })
            total_verificado += precio
        
        # Agrupar exámenes iguales
        examenes_agrupados = {}
        for item in examenes_con_precio:
            nombre = item['nombre']
            if nombre in examenes_agrupados:
                examenes_agrupados[nombre]['cantidad'] += 1
                examenes_agrupados[nombre]['subtotal'] += item['precio_unitario']
            else:
                examenes_agrupados[nombre] = {
                    'precio_unitario': item['precio_unitario'],
                    'cantidad': 1,
                    'subtotal': item['precio_unitario']
                }
        
        context = {
            'venta': venta,
            'fecha': venta.fecha_creacion,
            'usuario': request.user,
            'examenes_agrupados': examenes_agrupados,  # Ahora con precios
            'total_examenes': len(ordenes_examen),
            'total_verificado': total_verificado,
        }
        
    except VentaAmbulatoria.DoesNotExist:
        # Datos de ejemplo para desarrollo
        context = {
            'venta': {
                'id': venta_id,
                'paciente_nombre': 'PACIENTE EXTERNO',
                'paciente_dni': '87654321',
                'total': '185.50',
                'metodo_pago': 'EFECTIVO',
            },
            'fecha': timezone.now(),
            'usuario': request.user,
            'examenes_agrupados': {
                'HEMOGRAMA COMPLETO': {
                    'precio_unitario': Decimal('85.50'),
                    'cantidad': 1,
                    'subtotal': Decimal('85.50')
                },
                'GLUCOSA EN AYUNAS': {
                    'precio_unitario': Decimal('50.00'),
                    'cantidad': 2,
                    'subtotal': Decimal('100.00')
                },
            },
            'total_examenes': 3,
            'total_verificado': Decimal('185.50'),
        }
    
    return render(request, 'consultas/laboratorio/ticket_venta.html', context)

# ==============================================
# VISTAS PARA CONSULTAS (MANTENIDAS)
# ==============================================
@login_required
@grupo_requerido('doctores','laboratorio','administrador')
def paciente_tiene_consulta_hoy(paciente, tarifa, doctor=None):
    """
    Verifica si un paciente ya tiene una consulta registrada hoy
    para el mismo tipo de consulta y especialidad.
    """
    hoy = now().date()

    consultas_hoy = Consulta.objects.filter(
        paciente=paciente,
        fecha__date=hoy,
        estado=Consulta.EstadoConsulta.REGISTRADA
    )

    if not consultas_hoy.exists():
        return False

    for consulta_existente in consultas_hoy:
        if (tarifa.tipo_consulta == 'GENERAL' and
            consulta_existente.tarifa.tipo_consulta == 'GENERAL'):
            return True

        if (tarifa.tipo_consulta == 'ECOGRAFIA' and
            consulta_existente.tarifa.tipo_consulta == 'ECOGRAFIA'):
            return True

        if (tarifa.tipo_consulta == 'ESPECIALIDAD' and
            consulta_existente.tarifa.tipo_consulta == 'ESPECIALIDAD' and
            tarifa.especialidad == consulta_existente.tarifa.especialidad):
            return True

        if doctor and consulta_existente.doctor == doctor:
            return True

    return False
############
@login_required
@grupo_requerido('recepcion','administrador')
def crear_consulta(request, paciente_id):
    from django.utils.timezone import now
    from django.shortcuts import render, redirect, get_object_or_404
    from django.contrib import messages
    from django.http import JsonResponse

    from pacientes.models import Paciente
    from tarifas.models import TarifaConsulta
    from .models import Consulta, CatalogoEcografia, TopicoCatalogo
    from .forms import ConsultaForm
    from pagos.models import OrdenPago
    from doctores.models import Doctor

    paciente = get_object_or_404(Paciente, id=paciente_id)

    # Obtener datos para los combos
    tarifas = TarifaConsulta.objects.filter(activo=True)
    ecografias = CatalogoEcografia.objects.filter(activo=True)
    topicos = TopicoCatalogo.objects.filter(activo=True)

    # Función para validar consulta duplicada
    def paciente_tiene_consulta_hoy(paciente, tarifa, doctor):
        return Consulta.objects.filter(
            paciente=paciente,
            tarifa=tarifa,
            doctor=doctor,
            fecha__date=now().date(),
            estado=Consulta.EstadoConsulta.REGISTRADA
        ).exists()

    # Detectar tipo de consulta desde el POST
    es_topico = False
    es_lectura = False
    if request.method == 'POST':
        tarifa_id = request.POST.get('tarifa')
        if tarifa_id:
            try:
                tarifa_temp = TarifaConsulta.objects.get(id=tarifa_id)
                es_topico = (tarifa_temp.tipo_consulta == 'TOPICO')
                es_lectura = (tarifa_temp.tipo_consulta == 'LECTURA_RESULTADOS')
            except:
                pass

    if request.method == 'POST':
        # Crear formulario
        form = ConsultaForm(request.POST)

        # Hacer doctor no requerido para TÓPICO
        if es_topico:
            form.fields['doctor'].required = False
        # Para LECTURA, el doctor SÍ es requerido (mensaje claro)

        ecografia_id = request.POST.get('ecografia')
        topico_id = request.POST.get('topico')
        precio_input = request.POST.get('precio', '0')

        print("="*50)
        print(f"📝 POST recibido - tarifa: {request.POST.get('tarifa')}")
        print(f"📝 ecografia_id: {ecografia_id}")
        print(f"📝 topico_id: {topico_id}")
        print(f"📝 precio_input: {precio_input}")
        print(f"📝 es_topico: {es_topico}")
        print(f"📝 es_lectura: {es_lectura}")

        if form.is_valid():
            print("✅ Formulario válido")
            tarifa = form.cleaned_data['tarifa']
            doctor = form.cleaned_data.get('doctor')
            hoy = now().date()

            # ========================================
            # VALIDACIÓN 1: Consulta duplicada
            # ========================================
            existe = paciente_tiene_consulta_hoy(paciente, tarifa, doctor)
            if existe:
                if tarifa.tipo_consulta == 'GENERAL':
                    mensaje_error = 'El paciente ya tiene una consulta GENERAL registrada para hoy.'
                elif tarifa.tipo_consulta == 'ECOGRAFIA':
                    mensaje_error = 'El paciente ya tiene una ECOGRAFÍA registrada para hoy.'
                elif tarifa.tipo_consulta == 'ESPECIALIDAD':
                    mensaje_error = f'El paciente ya tiene una consulta de {tarifa.get_especialidad_display()} registrada para hoy.'
                else:
                    mensaje_error = 'El paciente ya tiene una consulta registrada para hoy.'

                return render(request, 'consultas/crear_consulta.html', {
                    'paciente': paciente,
                    'form': form,
                    'tarifas': tarifas,
                    'ecografias': ecografias,
                    'topicos': topicos,
                    'error': mensaje_error
                })

            # ========================================
            # VALIDACIÓN 2: Ecografía obligatoria
            # ========================================
            if tarifa.tipo_consulta == 'ECOGRAFIA' and not ecografia_id:
                return render(request, 'consultas/crear_consulta.html', {
                    'paciente': paciente,
                    'form': form,
                    'tarifas': tarifas,
                    'ecografias': ecografias,
                    'topicos': topicos,
                    'error': 'Debe seleccionar un tipo de ecografía.'
                })

            # ========================================
            # VALIDACIÓN 3: Tópico obligatorio
            # ========================================
            if tarifa.tipo_consulta == 'TOPICO' and not topico_id:
                return render(request, 'consultas/crear_consulta.html', {
                    'paciente': paciente,
                    'form': form,
                    'tarifas': tarifas,
                    'ecografias': ecografias,
                    'topicos': topicos,
                    'error': 'Debe seleccionar un procedimiento de tópico.'
                })

            # ========================================
            # VALIDACIÓN 4: Doctor duplicado (excepto tópico)
            # ========================================
            if tarifa.tipo_consulta != 'TOPICO' and doctor:
                existe_con_doctor = Consulta.objects.filter(
                    paciente=paciente,
                    doctor=doctor,
                    fecha__date=hoy,
                    estado=Consulta.EstadoConsulta.REGISTRADA
                ).exists()

                if existe_con_doctor:
                    return render(request, 'consultas/crear_consulta.html', {
                        'paciente': paciente,
                        'form': form,
                        'tarifas': tarifas,
                        'ecografias': ecografias,
                        'topicos': topicos,
                        'error': f'El paciente ya tiene una consulta registrada hoy con el Dr. {doctor.nombres} {doctor.apellidos}.'
                    })

            # ========================================
            # CREAR CONSULTA
            # ========================================
            consulta = form.save(commit=False)
            consulta.paciente = paciente

            # Variable para tracking del precio
            precio_final = 0

            # ASIGNAR SEGÚN TIPO DE CONSULTA
            if tarifa.tipo_consulta == 'ECOGRAFIA' and ecografia_id:
                ecografia = CatalogoEcografia.objects.get(id=ecografia_id)
                consulta.ecografia_catalogo = ecografia
                precio_final = ecografia.precio
                consulta.descripcion_servicio = ecografia.nombre
                consulta.topico_catalogo = None
                consulta.doctor = doctor
                print(f"💰 Ecografía - Precio catálogo: {ecografia.precio}")

            elif tarifa.tipo_consulta == 'TOPICO' and topico_id:
                topico = TopicoCatalogo.objects.get(id=topico_id)
                consulta.topico_catalogo = topico

                try:
                    precio_limpio = str(precio_input).replace(',', '.').strip()
                    if precio_limpio and precio_limpio not in ['0', '0.00', '']:
                        precio_final = float(precio_limpio)
                        print(f"💰 Tópico - Precio editado: {precio_final}")
                    else:
                        precio_final = topico.precio
                        print(f"💰 Tópico - Usando precio catálogo: {topico.precio}")
                except (ValueError, TypeError, AttributeError) as e:
                    print(f"⚠️ Error en precio: {e}, usando catálogo: {topico.precio}")
                    precio_final = topico.precio

                consulta.descripcion_servicio = f"{topico.nombre} (S/{precio_final})"
                consulta.ecografia_catalogo = None
                consulta.doctor = None
                print(f"💰 Precio final tópico: {precio_final}")

            elif tarifa.tipo_consulta == 'LECTURA_RESULTADOS':
                # LECTURA DE RESULTADOS
                precio_final = 0
                consulta.ecografia_catalogo = None
                consulta.topico_catalogo = None
                consulta.doctor = doctor  # Requerido, pero ya validado
                consulta.es_lectura_resultados = True
                print(f"📋 Lectura de resultados - Doctor: {doctor}")

            else:
                # GENERAL, ESPECIALIDAD
                precio_final = tarifa.precio
                consulta.ecografia_catalogo = None
                consulta.topico_catalogo = None
                consulta.doctor = doctor
                print(f"💰 Otro tipo - Precio tarifa: {tarifa.precio}")

            # ASIGNAR PRECIO
            consulta.precio = precio_final

            # Guardar consulta
            consulta.save()
            print(f"✅ Consulta #{consulta.id} guardada - Doctor: {consulta.doctor} - Precio: {consulta.precio}")

            # DOBLE VERIFICACIÓN DE PRECIO
            if float(consulta.precio) != float(precio_final):
                print(f"⚠️ Precio incorrecto en BD: {consulta.precio}, forzando a {precio_final}")
                Consulta.objects.filter(id=consulta.id).update(precio=precio_final)
                consulta.refresh_from_db()
                print(f"💰 Precio después de forzar: {consulta.precio}")

            # CREAR ORDEN DE PAGO
            orden = OrdenPago.objects.create(
                consulta=consulta,
                monto=consulta.precio,
                estado='PENDIENTE'
            )
            print(f"✅ Orden de pago #{orden.id} creada - Monto: {orden.monto}")
            print("="*50)

            messages.success(request, f'✅ Consulta #{consulta.id} - {tarifa.get_tipo_consulta_display()} - S/{consulta.precio}')
            return redirect('detalle_orden_pago', orden.id)

        else:
            # ❌ Formulario inválido - Mensajes personalizados
            print(f"❌ Formulario inválido: {form.errors}")
            
            # Personalizar mensaje de error
            error_message = "Por favor, corrija los errores en el formulario."
            
            if 'doctor' in form.errors:
                if 'This field is required' in str(form.errors['doctor']):
                    if es_lectura:
                        error_message = "👨‍⚕️ Para una consulta de LECTURA DE RESULTADOS debe seleccionar un médico que interpretará los resultados."
                    elif not es_topico:
                        error_message = "👨‍⚕️ Debe seleccionar un médico para esta consulta."
            
            return render(request, 'consultas/crear_consulta.html', {
                'paciente': paciente,
                'form': form,
                'tarifas': tarifas,
                'ecografias': ecografias,
                'topicos': topicos,
                'error': error_message
            })

    else:
        # GET request
        form = ConsultaForm()

    return render(request, 'consultas/crear_consulta.html', {
        'paciente': paciente,
        'form': form,
        'tarifas': tarifas,
        'ecografias': ecografias,
        'topicos': topicos,
    })
##################################################
# En views.py, añadir:
@login_required
@grupo_requerido('laboratorio', 'administrador')
def buscar_ordenes_laboratorio(request):
    """Búsqueda global de órdenes por DNI en todos los estados"""
    dni_buscado = request.GET.get('dni', '').strip()
    estado_filtro = request.GET.get('estado', 'TODOS')

    if not dni_buscado:
        # Redirigir al dashboard si no hay búsqueda
        return redirect('lista_ordenes_pendientes')

    # Filtrar por DNI en TODOS los estados
    ordenes = OrdenExamen.objects.filter(
        tipo_examen='LABORATORIO'
    ).select_related(
        'consulta__paciente',
        'venta_ambulatoria',
        'tecnico_asignado'
    )

    # Aplicar filtro DNI
    dni_limpio = dni_buscado.replace(' ', '').replace('-', '').replace('.', '')

    if dni_limpio.isdigit() and len(dni_limpio) == 8:
        ordenes = ordenes.filter(
            Q(consulta__paciente__dni=dni_limpio) |
            Q(venta_ambulatoria__paciente_dni=dni_limpio)
        )
    else:
        ordenes = ordenes.filter(
            Q(consulta__paciente__dni__icontains=dni_buscado) |
            Q(venta_ambulatoria__paciente_dni__icontains=dni_buscado)
        )

    # Filtrar por estado si se especifica
    if estado_filtro != 'TODOS':
        ordenes = ordenes.filter(estado=estado_filtro)

    # Ordenar por fecha
    ordenes = ordenes.order_by('-fecha_solicitud')

    # Contar por estado
    contadores = {
        'SOLICITADO': ordenes.filter(estado='SOLICITADO').count(),
        'EN_PROCESO': ordenes.filter(estado='EN_PROCESO').count(),
        'COMPLETADO': ordenes.filter(estado='COMPLETADO').count(),
        'ENTREGADO': ordenes.filter(estado='ENTREGADO').count(),
        'TODOS': ordenes.count(),
    }

    context = {
        'ordenes': ordenes,
        'dni_buscado': dni_buscado,
        'estado_filtro': estado_filtro,
        'contadores': contadores,
        'titulo': f'Búsqueda: {dni_buscado}',
    }

    return render(request, 'consultas/laboratorio/busqueda_global.html', context)


#############################
@login_required
@grupo_requerido('laboratorio')
def buscar_ordenes_simple(request):
    """Búsqueda simple por DNI"""
    dni = request.GET.get('dni', '').strip()

    if not dni:
        return redirect('lista_ordenes_pendientes')

    # Buscar simple
    ordenes = OrdenExamen.objects.filter(
        Q(consulta__paciente__dni__icontains=dni) |
        Q(venta_ambulatoria__paciente_dni__icontains=dni),
        tipo_examen='LABORATORIO'
    ).order_by('-fecha_solicitud')

    context = {
        'ordenes': ordenes,
        'dni_buscado': dni,
        'titulo': f'Búsqueda: {dni}',
    }

    return render(request, 'consultas/laboratorio/busqueda_simple.html', context)


########################
@login_required
@grupo_requerido('laboratorio', 'administrador')
def buscar_ordenes_global(request):
    """Búsqueda global de órdenes por DNI"""
    dni_buscado = request.GET.get('dni', '').strip()

    if not dni_buscado:
        # Redirigir a pendientes si no hay búsqueda
        return redirect('lista_ordenes_pendientes')

    # Buscar en TODOS los estados
    ordenes = OrdenExamen.objects.filter(
        tipo_examen='LABORATORIO'
    ).select_related(
        'consulta__paciente',
        'venta_ambulatoria'
    )

    # Filtrar por DNI
    if dni_buscado:
        ordenes = ordenes.filter(
            Q(consulta__paciente__dni__icontains=dni_buscado) |
            Q(venta_ambulatoria__paciente_dni__icontains=dni_buscado)
        )

    # Ordenar por fecha más reciente
    ordenes = ordenes.order_by('-fecha_solicitud')

    # Contar por estado
    contadores = {
        'pendientes': ordenes.filter(estado='SOLICITADO').count(),
        'en_proceso': ordenes.filter(estado='EN_PROCESO').count(),
        'completados': ordenes.filter(estado='COMPLETADO').count(),
        'entregados': ordenes.filter(estado='ENTREGADO').count(),
        'total': ordenes.count(),
    }

    context = {
        'ordenes': ordenes,
        'dni_buscado': dni_buscado,
        'contadores': contadores,
        'titulo': f'Búsqueda: {dni_buscado}',
    }

    return render(request, 'consultas/laboratorio/busqueda_global.html', context)


########
@login_required
@require_POST  # IMPORTANTE: Solo por POST para seguridad
def cancelar_eliminar_orden(request, orden_id):
    """
    Elimina una orden pendiente del sistema
    Único caso: paciente no se presentó para los análisis
    """
    try:
        orden = get_object_or_404(OrdenExamen, id=orden_id)

        # SOLO se puede eliminar si cumple estas condiciones:
        if orden.estado != 'SOLICITADO':
            messages.error(request, f'No se puede eliminar la orden #{orden_id}. Solo órdenes pendientes.')
            return redirect('lista_ordenes_pendientes')

        if orden.tipo_examen != 'LABORATORIO':
            messages.error(request, f'La orden #{orden_id} no es de laboratorio.')
            return redirect('lista_ordenes_pendientes')

        if orden.tecnico_asignado:
            messages.error(request, f'No se puede eliminar la orden #{orden_id} porque ya está asignada.')
            return redirect('lista_ordenes_pendientes')

        # Guardar info para el mensaje (antes de eliminar)
        paciente_nombre = ""
        if orden.consulta:
            paciente_nombre = f"{orden.consulta.paciente.nombres} {orden.consulta.paciente.apellidos}"
        elif orden.venta_ambulatoria:
            paciente_nombre = orden.venta_ambulatoria.paciente_nombre

        # ELIMINAR la orden (no cancelar, ELIMINAR)
        orden.delete()

        messages.success(request,
            f'✅ Orden #{orden_id} eliminada correctamente. '
            f'Paciente: {paciente_nombre}'
        )

    except Exception as e:
        messages.error(request, f'❌ Error al eliminar orden: {str(e)}')

    return redirect('lista_ordenes_pendientes')



#################
@login_required
@grupo_requerido('laboratorio', 'administrador')
@require_POST
def cobrar_ordenes_seleccionadas(request):
    """Versión simplificada y segura - CON VALIDACIÓN DE NO MEZCLA"""

    # ✅ VALIDACIÓN SIMPLE: Si tiene exámenes nuevos, NO agregar órdenes internas
    carrito = request.session.get('carrito_laboratorio', {})
    tiene_examenes_nuevos = any(not k.startswith('orden_interna_') for k in carrito.keys())

    if tiene_examenes_nuevos:
        messages.error(request, '❌ No puedes agregar órdenes internas porque ya tienes exámenes nuevos en el carrito.')
        return redirect('lista_ordenes_pendientes')

    # RESTO DE TU CÓDIGO ORIGINAL (con corrección de precios)
    ordenes_ids = [int(id) for id in request.POST.getlist('ordenes_ids') if id.isdigit()]

    if not ordenes_ids:
        messages.error(request, '❌ No se seleccionaron órdenes')
        return redirect('lista_ordenes_pendientes')

    # Inicializar carrito
    request.session.setdefault('carrito_laboratorio', {})
    carrito = request.session['carrito_laboratorio']

    # Contar items antes
    items_iniciales = len(carrito)

    for orden_id in ordenes_ids:
        try:
            orden = OrdenExamen.objects.get(id=orden_id)

            # Solo órdenes internas no pagadas
            if orden.consulta and not orden.pagado:
                item_id = f"orden_interna_{orden.id}"

                if item_id not in carrito:
                    # ✅ CORRECCIÓN: BUSCAR PRECIO REAL EN CATÁLOGO
                    precio_final = "0.00"
                    
                    # BUSCAR EN CATÁLOGO SEGÚN TIPO DE EXAMEN
                    if orden.tipo_examen == 'LABORATORIO':
                        # Buscar en CatalogoExamen
                        try:
                            examen_catalogo = CatalogoExamen.objects.get(
                                nombre=orden.examen_especifico,
                                tipo='LABORATORIO',
                                activo=True
                            )
                            precio_final = str(examen_catalogo.precio)
                        except CatalogoExamen.DoesNotExist:
                            # Si no encuentra, usar 50 como fallback (igual que antes)
                            precio_final = "50.00"
                            
                    elif orden.tipo_examen == 'ECOGRAFIA':
                        # Buscar en CatalogoEcografia
                        try:
                            eco_catalogo = CatalogoEcografia.objects.get(
                                nombre=orden.examen_especifico,
                                activo=True
                            )
                            precio_final = str(eco_catalogo.precio)
                        except CatalogoEcografia.DoesNotExist:
                            precio_final = "50.00"
                    else:
                        # Para otros tipos, usar 50
                        precio_final = "50.00"

                    carrito[item_id] = {
                        'id': item_id,
                        'orden_id': orden.id,
                        'nombre': f"Orden #{orden.id} - {orden.examen_especifico}",
                        'tipo': 'INTERNO_POR_COBRAR',
                        'precio': precio_final,  # ✅ AHORA PRECIO REAL
                        'cantidad': 1,
                        'paciente': orden.get_paciente_nombre(),
                        'es_interno': True,
                        # ✅ NUEVO: Guardar info del catálogo para referencia
                        'tipo_examen': orden.tipo_examen,
                        'examen_especifico': orden.examen_especifico,
                    }

        except (OrdenExamen.DoesNotExist, ValueError):
            continue

    request.session.modified = True

    items_finales = len(carrito)
    items_agregados = items_finales - items_iniciales

    if items_agregados > 0:
        messages.success(request, f'✅ {items_agregados} orden(es) agregada(s) al carrito')
        return redirect('punto_venta_laboratorio')
    else:
        messages.error(request, '❌ No se agregaron nuevas órdenes al carrito')
        return redirect('lista_ordenes_pendientes')
##############################
@login_required
@grupo_requerido('laboratorio', 'administrador')
def ticket_cobro_interno(request, venta_id):
    """Ticket para cobro de órdenes internas - VERSIÓN CORREGIDA CON DESCUENTO"""
    venta = get_object_or_404(VentaAmbulatoria, id=venta_id)

    # Obtener órdenes cobradas de la sesión
    datos_cobro = request.session.get('ultimas_ordenes_cobradas', {})
    ordenes_detalles = datos_cobro.get('detalles', [])

    # DEBUG: Ver qué datos hay en la sesión
    print(f"[DEBUG TICKET] Datos en sesión: {datos_cobro}")
    print(f"[DEBUG TICKET] Descuento en sesión: {datos_cobro.get('descuento', 'NO HAY')}")
    print(f"[DEBUG TICKET] Total en sesión: {datos_cobro.get('total', 'NO HAY')}")

    # Convertir los precios de float (de JSON) a Decimal para el template
    ordenes_con_precio_correcto = []
    total_sin_descuento = Decimal('0.00')
    total_con_descuento = Decimal('0.00')

    if ordenes_detalles:
        for oc in ordenes_detalles:
            try:
                # Precio original (desde datos de sesión)
                precio_original = Decimal(str(oc.get('precio_original', 50.00)))
            except:
                precio_original = Decimal('50.00')

            try:
                # Precio con descuento (desde datos de sesión)
                precio_final = Decimal(str(oc.get('precio_final', oc.get('precio_con_descuento', oc.get('precio', precio_original)))))
            except:
                precio_final = precio_original

            descuento_individual = precio_original - precio_final

            ordenes_con_precio_correcto.append({
                'id': oc.get('id'),
                'examen': oc.get('examen', 'Examen no especificado'),
                'precio_original': precio_original,
                'precio_final': precio_final,
                'descuento_individual': descuento_individual,
                'paciente': oc.get('paciente', 'Paciente no identificado'),
                'dni': oc.get('dni', '')
            })

            total_sin_descuento += precio_original
            total_con_descuento += precio_final

    # Si no hay datos en sesión, buscar en BD (fallback)
    if not ordenes_con_precio_correcto:
        # Buscar órdenes pagadas recientemente (últimos 5 minutos)
        desde = timezone.now() - timedelta(minutes=5)

        ordenes_db = OrdenExamen.objects.filter(
            pagado=True,
            fecha_pago__gte=desde,
            consulta__isnull=False  # Solo internas
        ).select_related('consulta__paciente').order_by('-fecha_pago')

        for orden in ordenes_db:
            # Precio original (asumir 50.00)
            precio_original = Decimal('50.00')

            # Precio final (con descuento si aplica)
            precio_final = orden.monto_pagado if hasattr(orden, 'monto_pagado') and orden.monto_pagado else precio_original

            descuento_individual = precio_original - precio_final

            ordenes_con_precio_correcto.append({
                'id': orden.id,
                'examen': orden.examen_especifico,
                'precio_original': precio_original,
                'precio_final': precio_final,
                'descuento_individual': descuento_individual,
                'paciente': orden.get_paciente_nombre(),
                'dni': orden.get_paciente_dni()
            })

            total_sin_descuento += precio_original
            total_con_descuento += precio_final

    # Obtener descuento de la sesión (¡USAR EL DE LA SESIÓN!)
    try:
        descuento_total = Decimal(str(datos_cobro.get('descuento', 0)))
    except:
        descuento_total = Decimal('0.00')

    # Si el descuento de la sesión es 0 pero venta tiene descuento, usar el de venta
    if descuento_total == 0 and venta.descuento_aplicado > 0:
        descuento_total = Decimal(str(venta.descuento_aplicado))
        print(f"[DEBUG TICKET] Usando descuento de venta: {descuento_total}")

    # Si todavía es 0, calcularlo a partir de los precios
    if descuento_total == 0:
        descuento_total = total_sin_descuento - total_con_descuento
        print(f"[DEBUG TICKET] Calculando descuento: {descuento_total}")

    # Calcular total final (usar el de la sesión si existe)
    try:
        total_final_sesion = Decimal(str(datos_cobro.get('total_con_descuento', total_con_descuento)))
        total_final = total_final_sesion
    except:
        total_final = total_con_descuento

    # DEBUG: Imprimir valores para diagnóstico
    print(f"[DEBUG TICKET] total_sin_descuento: {total_sin_descuento}")
    print(f"[DEBUG TICKET] total_con_descuento: {total_con_descuento}")
    print(f"[DEBUG TICKET] descuento_total: {descuento_total}")
    print(f"[DEBUG TICKET] total_final: {total_final}")
    print(f"[DEBUG TICKET] Descuento de venta: {venta.descuento_aplicado}")

    # Preparar contexto
    context = {
        'venta': venta,
        'ordenes': ordenes_con_precio_correcto,
        'total_sin_descuento': total_sin_descuento,
        'descuento_total': descuento_total,
        'total_final': total_final,
        'metodo_pago': datos_cobro.get('metodo_pago', 'EFECTIVO'),
        'fecha': timezone.now(),
        'usuario': request.user,
        'es_cobro_interno': True,
        'cantidad_ordenes': len(ordenes_con_precio_correcto),
    }

    # Limpiar datos de sesión después de usarlos
    if 'ultimas_ordenes_cobradas' in request.session:
        del request.session['ultimas_ordenes_cobradas']
        request.session.modified = True

    return render(request, 'consultas/laboratorio/ticket_cobro_interno.html', context)

##################
# /home/luis/policlinico/consultas/api_views.py
# /home/luis/policlinico/consultas/api_views.py

from django.http import JsonResponse
from doctores.models import Doctor
import traceback

# /home/luis/policlinico/consultas/views.py
@login_required
@grupo_requerido('recepcion','administrador')
def doctores_por_especialidad(request):
    """
    API endpoint para obtener doctores filtrados por especialidad
    VERSIÓN SIMPLE - SIN ERRORES
    """
    try:
        especialidad = request.GET.get('especialidad')
        cargar_todos = request.GET.get('todos') == 'true'

        print(f"🔍 API - especialidad: {especialidad}, todos: {cargar_todos}")

        # Versión simple - sin filtros complejos
        if cargar_todos:
            doctores = Doctor.objects.filter(activo=True)
        elif especialidad:
            # ✅ FILTRO SIMPLE - sin distinct ni relaciones complejas
            doctores = Doctor.objects.filter(activo=True)
            # Filtramos en Python en lugar de SQL
            resultado = []
            for d in doctores:
                for e in d.especialidades.all():
                    if e.nombre == especialidad:
                        resultado.append(d)
                        break
            doctores = resultado
        else:
            doctores = []

        data = []
        for d in doctores:
            # Obtener especialidades de forma segura
            especialidades = []
            for e in d.especialidades.all():
                try:
                    especialidades.append(e.get_nombre_display())
                except:
                    especialidades.append(e.nombre)

            data.append({
                'id': d.id,
                'nombre_completo': f"Dr. {d.nombres} {d.apellidos}",
                'especialidades': ", ".join(especialidades)
            })

        return JsonResponse(data, safe=False)

    except Exception as e:
        print(f"❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        # Devolver array vacío en lugar de error
        return JsonResponse([], safe=False)


#############
# En views.py, actualiza el mapa de formatos:
# ============================================
# MAPA DE FORMATOS POR TIPO DE EXAMEN (ACTUALIZADO)
# ============================================
FORMATOS_EXAMENES = {
    # ===== FORMATO 1: HEMOGRAMA =====
    'HEMOGRAMA COMPLETO': 'consultas/laboratorio/formatos/BHCG-Cuantitativo-new.html',
    'HEMOGLOBINA': 'consultas/laboratorio/formatos/hemograma.html',
    'RECUENTO DE PLAQUETAS': 'consultas/laboratorio/formatos/hemograma.html',
    'RECUENTO DE GLÓBULOS ROJOS': 'consultas/laboratorio/formatos/hemograma.html',
    'CONSTANTES CORPUSCULARES': 'consultas/laboratorio/formatos/hemograma.html',
    'VELOCIDAD DE SEDIMENTACIÓN GLOBULAR': 'consultas/laboratorio/formatos/hemograma.html',
    'HEMATOCRITO': 'consultas/laboratorio/formatos/hemograma.html',
    'LEUCOCITOS': 'consultas/laboratorio/formatos/hemograma.html',

    # ===== FORMATO 2: PERFIL RENAL =====
    'PERFIL RENAL': 'consultas/laboratorio/formatos/perfil_renal.html',
    'ÁCIDO ÚRICO': 'consultas/laboratorio/formatos/perfil_renal.html',
    'UREA': 'consultas/laboratorio/formatos/perfil_renal.html',
    'CREATININA': 'consultas/laboratorio/formatos/perfil_renal.html',
    'DEPURACIÓN CREATININA': 'consultas/laboratorio/formatos/perfil_renal.html',
    'CALCIO ORINA 24H': 'consultas/laboratorio/formatos/perfil_renal.html',
    'MICROALBUMINURIA': 'consultas/laboratorio/formatos/perfil_renal.html',

    # ===== FORMATO 3: PERFIL HEPÁTICO =====
    'PERFIL HEPÁTICO': 'consultas/laboratorio/formatos/perfil_hepatico.html',
    'TGO': 'consultas/laboratorio/formatos/perfil_hepatico.html',
    'TGP': 'consultas/laboratorio/formatos/perfil_hepatico.html',
    'FOSFATASA ALCALINA': 'consultas/laboratorio/formatos/perfil_hepatico.html',
    'BILIRRUBINA TOTAL Y FRACCIONADA': 'consultas/laboratorio/formatos/perfil_hepatico.html',
    'PROTEÍNAS TOTAL Y FRACCIONADAS': 'consultas/laboratorio/formatos/perfil_hepatico.html',
    'GGTP': 'consultas/laboratorio/formatos/perfil_hepatico.html',
    'LDH': 'consultas/laboratorio/formatos/perfil_hepatico.html',
    'FERRITINA': 'consultas/laboratorio/formatos/perfil_hepatico.html',

    # ===== FORMATO 4: PERFIL LIPÍDICO =====
    'PERFIL LIPÍDICO': 'consultas/laboratorio/formatos/perfil_lipidico.html',
    'COLESTEROL TOTAL': 'consultas/laboratorio/formatos/perfil_lipidico.html',
    'COLESTEROL HDL': 'consultas/laboratorio/formatos/perfil_lipidico.html',
    'COLESTEROL LDL': 'consultas/laboratorio/formatos/perfil_lipidico.html',
    'COLESTEROL VLDL': 'consultas/laboratorio/formatos/perfil_lipidico.html',
    'TRIGLICÉRIDOS': 'consultas/laboratorio/formatos/perfil_lipidico.html',

    # ===== FORMATO 5: HORMONAS (VALOR ÚNICO) =====
    'TSH': 'consultas/laboratorio/formatos/hormonas.html',
    'T3': 'consultas/laboratorio/formatos/hormonas.html',
    'T4': 'consultas/laboratorio/formatos/hormonas.html',
    'T4 LIBRE': 'consultas/laboratorio/formatos/hormonas.html',
    'FSH': 'consultas/laboratorio/formatos/hormonas.html',
    'LH': 'consultas/laboratorio/formatos/hormonas.html',
    'PROLACTINA': 'consultas/laboratorio/formatos/hormonas.html',
    'ESTRADIOL': 'consultas/laboratorio/formatos/hormonas.html',
    'INSULINA': 'consultas/laboratorio/formatos/hormonas.html',
    'IGE': 'consultas/laboratorio/formatos/hormonas.html',
    'HORMONA ANTIMULLERIANA': 'consultas/laboratorio/formatos/hormonas.html',
    'BHCG CUANTITATIVO': 'consultas/laboratorio/formatos/hormonas.html',

    # ===== FORMATO 6: PSA (ESPECIAL) =====
    'PSA TOTAL': 'consultas/laboratorio/formatos/psa.html',
    'PSA LIBRE': 'consultas/laboratorio/formatos/psa.html',

    # ===== FORMATO 7: METABOLISMO =====
    'GLUCOSA': 'consultas/laboratorio/formatos/metabolismo.html',
    'HEMOGLOBINA GLICOSILADA': 'consultas/laboratorio/formatos/metabolismo.html',
    'TOLERANCIA A LA GLUCOSA': 'consultas/laboratorio/formatos/metabolismo.html',

    # ===== FORMATO 8: EXÁMENES DE ORINA =====
    'EXAMEN COMPLETO DE ORINA': 'consultas/laboratorio/formatos/examen_orina.html',
    'UROCULTIVO + ANTIBIOGRAMA': 'consultas/laboratorio/formatos/examen_orina.html',
    'PROTEINURIA CUALITATIVA': 'consultas/laboratorio/formatos/examen_orina.html',
    'PROTEINURIA ORINA 24H': 'consultas/laboratorio/formatos/examen_orina.html',
    'MICROALBUMINURIA ORINA 24H': 'consultas/laboratorio/formatos/examen_orina.html',

    # ===== FORMATO 9: EXÁMENES DE HECES =====
    'EXAMEN DIRECTO DE HECES': 'consultas/laboratorio/formatos/examen_heces.html',
    'EXAMEN SERIADO DE HECES': 'consultas/laboratorio/formatos/examen_heces.html',
    'SANGRE OCULTA EN HECES': 'consultas/laboratorio/formatos/examen_heces.html',
    'LEUCOCITOS EN HECES': 'consultas/laboratorio/formatos/examen_heces.html',
    'REACCIÓN INFLAMATORIA': 'consultas/laboratorio/formatos/examen_heces.html',
    'TEST DE GRAHAM': 'consultas/laboratorio/formatos/examen_heces.html',

    # ===== FORMATO 10: SEROLOGÍA =====
    'VIH 1 Y 2': 'consultas/laboratorio/formatos/serologia.html',
    'RPR (SÍFILIS)': 'consultas/laboratorio/formatos/serologia.html',
    'AGSHB (HEPATITIS B)': 'consultas/laboratorio/formatos/serologia.html',
    'VHC (HEPATITIS C)': 'consultas/laboratorio/formatos/serologia.html',
    'IGM HA (HEPATITIS A)': 'consultas/laboratorio/formatos/serologia.html',
    'IGM HELICOBACTER PYLORI': 'consultas/laboratorio/formatos/serologia.html',
    'REACCIÓN DE WIDAL': 'consultas/laboratorio/formatos/serologia.html',

    # ===== FORMATO 11: MARCADORES INFLAMATORIOS =====
    'PROTEÍNA C REACTIVA (PCR)': 'consultas/laboratorio/formatos/inflamatorios.html',
    'PCR CUALITATIVO': 'consultas/laboratorio/formatos/inflamatorios.html',
    'FACTOR REUMATOIDEO LÁTEX': 'consultas/laboratorio/formatos/inflamatorios.html',
    'ASO': 'consultas/laboratorio/formatos/inflamatorios.html',

    # ===== FORMATO 12: COAGULACIÓN =====
    'TIEMPO DE COAGULACIÓN': 'consultas/laboratorio/formatos/coagulacion.html',
    'TIEMPO DE SANGRÍA': 'consultas/laboratorio/formatos/coagulacion.html',
    'TIEMPO DE PROTROMBINA + INR': 'consultas/laboratorio/formatos/coagulacion.html',
    'TIEMPO PARCIAL DE TROMBOPLASTINA': 'consultas/laboratorio/formatos/coagulacion.html',

    # ===== FORMATO 13: MICROBIOLOGÍA =====
    'RASPADO PIEL KOH': 'consultas/laboratorio/formatos/microbiologia.html',
    'SECRECIÓN VAGINAL': 'consultas/laboratorio/formatos/microbiologia.html',
    'SECRECIÓN URETRAL': 'consultas/laboratorio/formatos/microbiologia.html',
    'COLORACIÓN DE GRAM': 'consultas/laboratorio/formatos/microbiologia.html',

    # ===== FORMATO 14: EMBARAZO =====
    'PREGNOSTICON': 'consultas/laboratorio/formatos/embarazo.html',

    # ===== FORMATO 15: BACILOSCOPÍA =====
    'BACILOSCOPÍA EN ESPUTO': 'consultas/laboratorio/formatos/baciloscopia.html',

    # ===== POR DEFECTO (si no encuentra) =====
    'DEFAULT': 'consultas/laboratorio/formatos/generico.html',
}

@login_required
@grupo_requerido('laboratorio', 'administrador')
def ingresar_resultados(request, examen_id):
    """
    Vista para ingresar resultados usando el formato específico del examen
    """
    examen = get_object_or_404(OrdenExamen, id=examen_id, estado='EN_PROCESO')
    
    # Verificar que esté asignado al laboratorista actual
    if examen.tecnico_asignado != request.user:
        messages.error(request, '❌ Este examen no está asignado a usted')
        return redirect('mis_ordenes_en_proceso')
    
    # Detectar qué plantilla usar según el nombre del examen
    nombre_examen = examen.examen_especifico.upper().strip()
    template_name = FORMATOS_EXAMENES.get(nombre_examen, FORMATOS_EXAMENES['DEFAULT'])
    
    if request.method == 'POST':
        # Recoger todos los campos del formulario
        resultados = {}
        for key, value in request.POST.items():
            if key not in ['csrfmiddlewaretoken', 'observaciones', 'btn_guardar']:
                if value.strip():
                    resultados[key] = value.strip()
        
        # Guardar observaciones si existen
        if request.POST.get('observaciones'):
            resultados['observaciones'] = request.POST.get('observaciones')
        
        # Guardar en JSON
        examen.guardar_resultados(resultados)
        
        messages.success(request, f'✅ Resultados de {examen.examen_especifico} guardados correctamente')
        return redirect('detalle_orden_examen', examen_id)
    
    # Mostrar el formulario
    context = {
        'orden': examen,
        'resultados_previos': examen.get_resultados_formateados(),
    }
    return render(request, template_name, context)

################# imprimir resultados
@login_required
@grupo_requerido('doctores', 'laboratorio', 'administrador', 'recepcion')
def imprimir_resultado(request, examen_id):
    """
    Vista para imprimir resultado de examen
    """
    examen = get_object_or_404(OrdenExamen, id=examen_id)

    # Verificar permisos para doctores
    if request.user.groups.filter(name='doctores').exists():
        if examen.consulta and examen.consulta.doctor != request.user.doctor:
            return HttpResponseForbidden("No tiene permiso para ver este resultado")

    # ============================================
    # 1. DETERMINAR QUÉ TEMPLATE DE IMPRESIÓN USAR
    # ============================================
    nombre_examen = examen.examen_especifico.upper().strip()
    
    if 'BHCG' in nombre_examen or 'BETA HCG' in nombre_examen:
        template_name = 'consultas/laboratorio/formatos/imprimir_bhcg.html'
    else:
        template_name = 'consultas/laboratorio/formatos/imprimir_generico.html'

    # ============================================
    # 2. OBTENER DATOS DEL PACIENTE (INCLUYENDO EDAD)
    # ============================================
    from datetime import date
    
    paciente_nombre = examen.get_paciente_nombre()
    paciente_dni = examen.get_paciente_dni()
    paciente_edad = "-"
    
    # Obtener paciente según el tipo de orden
    if examen.consulta:
        paciente = examen.consulta.paciente
    else:
        paciente = None
    
    # Calcular edad si existe fecha de nacimiento
    if paciente and hasattr(paciente, 'fecha_nacimiento') and paciente.fecha_nacimiento:
        hoy = date.today()
        edad = hoy.year - paciente.fecha_nacimiento.year
        if hoy.month < paciente.fecha_nacimiento.month or (hoy.month == paciente.fecha_nacimiento.month and hoy.day < paciente.fecha_nacimiento.day):
            edad -= 1
        paciente_edad = edad
    
    # ============================================
    # 3. PREPARAR RESULTADOS
    # ============================================
    resultados = examen.get_resultados_formateados()
    
    # ============================================
    # 4. CONTEXTO PARA EL TEMPLATE
    # ============================================
    context = {
        'orden': examen,
        'resultados': resultados,
        'paciente_nombre': paciente_nombre,
        'paciente_dni': paciente_dni,
        'paciente_edad': paciente_edad,
        'tecnico_nombre': examen.tecnico_asignado.get_full_name() if examen.tecnico_asignado else "-",
    }

    return render(request, template_name, context)

############ administrador ###############
@login_required
@grupo_requerido('administrador')
def api_consultas_dia(request):
    """API para obtener consultas del día en formato JSON"""
    from datetime import datetime, date
    from .models import Consulta
    
    fecha_str = request.GET.get('fecha', '')
    doctor_id = request.GET.get('doctor', 'todos')
    
    try:
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date() if fecha_str else date.today()
    except:
        fecha = date.today()
    
    consultas = Consulta.objects.filter(
        fecha__date=fecha
    ).select_related('paciente', 'doctor__usuario').order_by('fecha')
    
    if doctor_id != 'todos':
        consultas = consultas.filter(doctor_id=doctor_id)
    
    data = []
    for c in consultas:
        data.append({
            'id': c.id,
            'hora': c.fecha.strftime('%H:%M'),
            'paciente': f"{c.paciente.nombres} {c.paciente.apellidos}",
            'dni': c.paciente.dni,
            'doctor': f"Dr. {c.doctor.nombres} {c.doctor.apellidos}" if c.doctor else 'No asignado',
            'tipo': c.tipo_consulta,  # ✅ Campo directo
            'estado': c.estado,
            'estado_display': c.get_estado_display(),  # ✅ Método correcto
        })
    
    return JsonResponse({'consultas': data})

################################
@login_required
@grupo_requerido('administrador')
def reportes_view(request):
    """Vista principal de reportes"""
    
    tipo_reporte = request.GET.get('tipo', 'consultas_periodo')
    fecha_desde = request.GET.get('desde', date.today().strftime('%Y-%m-%d'))
    fecha_hasta = request.GET.get('hasta', date.today().strftime('%Y-%m-%d'))
    page = request.GET.get('page', 1)
    
    try:
        desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    except:
        desde = date.today()
        hasta = date.today()
    
    # Diccionario de funciones
    reportes_func = {
        'consultas_periodo': get_consultas_periodo,
        'consultas_doctor': get_consultas_doctor,
        'ingresos_consultas': get_ingresos_consultas,
        'laboratorio': get_laboratorio,
        'farmacia': get_farmacia,
        'ingresos_totales': get_ingresos_totales,
        'metodos_pago': get_metodos_pago,
        'inventario': get_inventario,
        'topico': get_topico,
        'compras': get_compras,
        'digemid': get_reporte_digemid,
    }
    
    # Llamar a la función correspondiente
    if tipo_reporte in reportes_func:
        if tipo_reporte == 'inventario':
            data = reportes_func[tipo_reporte](page)
        elif tipo_reporte == 'compras':
            data = reportes_func[tipo_reporte](desde, hasta, page, request)
        else:
            data = reportes_func[tipo_reporte](desde, hasta, page)
    else:
        data = {'mensaje': 'Reporte no encontrado'}
    
    context = {
        'tipo_reporte': tipo_reporte,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'data': data,
        'reportes': [
            {'value': 'consultas_periodo', 'label': '📊 Consultas por período'},
            {'value': 'consultas_doctor', 'label': '👨‍⚕️ Consultas por doctor'},
            {'value': 'laboratorio', 'label': '🧪 Laboratorio'},
            {'value': 'farmacia', 'label': '💊 Farmacia'},
            {'value': 'ingresos_totales', 'label': '💵 Ingresos totales'},
            {'value': 'metodos_pago', 'label': '💳 Métodos de pago'},
            {'value': 'inventario', 'label': '📦 Inventario'},
            {'value': 'topico', 'label': '🩺 Tópico (Mano de obra)'},
            {'value': 'compras', 'label': '💰 Compras a proveedores'},
            {'value': 'digemid', 'label': '📋 Reporte DIGEMID (Precios)'},
        ]
    }
    return render(request, 'reportes.html', context)
#####################

def get_consultas_periodo(desde, hasta, page=1):
    """Reporte de consultas por período con paginación"""
    from .models import Consulta
    from pagos.models import OrdenPago
    from django.db.models import Sum
    
    # Consultas en el período
    consultas = Consulta.objects.filter(
        fecha__date__gte=desde,
        fecha__date__lte=hasta
    ).select_related('paciente', 'doctor').order_by('-fecha')
    
    total_consultas = consultas.count()
    atendidas = consultas.filter(estado='ATENDIDA').count()
    
    # Ingresos por método de pago
    pagos = OrdenPago.objects.filter(
        consulta__in=consultas,
        estado='PAGADO'
    )
    
    efectivo = pagos.filter(metodo_pago='EFECTIVO').aggregate(total=Sum('monto'))['total'] or 0
    yape = pagos.filter(metodo_pago='YAPE').aggregate(total=Sum('monto'))['total'] or 0
    plin = pagos.filter(metodo_pago='PLIN').aggregate(total=Sum('monto'))['total'] or 0
    total_ingresos = efectivo + yape + plin
    
    # PAGINACIÓN
    paginator = Paginator(consultas, 20)  # 20 items por página
    try:
        consultas_page = paginator.page(page)
    except PageNotAnInteger:
        consultas_page = paginator.page(1)
    except EmptyPage:
        consultas_page = paginator.page(paginator.num_pages)
    
    # Listado detallado de la página actual
    detalle = []
    for c in consultas_page:
        detalle.append({
            'fecha': c.fecha.strftime('%d/%m/%Y'),
            'hora': c.fecha.strftime('%H:%M'),
            'paciente': f"{c.paciente.nombres} {c.paciente.apellidos}",
            'doctor': f"Dr. {c.doctor.nombres} {c.doctor.apellidos}" if c.doctor else 'No asignado',
            'tipo': c.tipo_consulta,
            'estado': c.get_estado_display(),
        })
    
    return {
        'total_consultas': total_consultas,
        'atendidas': atendidas,
        'efectivo': efectivo,
        'yape': yape,
        'plin': plin,
        'total_ingresos': total_ingresos,
        'detalle': detalle,
        'total_registros': total_consultas,
        'paginator': paginator,
        'page_obj': consultas_page,
        'page': page,
    }

#########
def get_consultas_doctor(desde, hasta, page=1):
    """Reporte de consultas agrupadas por doctor y tipo de consulta"""
    from django.db.models import Count, Sum
    from django.core.paginator import Paginator
    from consultas.models import Consulta
    from doctores.models import Doctor
    
    # Obtener todas las consultas en el rango
    consultas = Consulta.objects.filter(
        fecha__date__gte=desde,
        fecha__date__lte=hasta
    ).select_related('doctor')
    
    # Definir tipos de consulta
    tipos = ['GENERAL', 'ESPECIALIDAD', 'ECOGRAFIA', 'LECTURA_RESULTADOS', 'TOPICO']
    
    # Obtener todos los doctores activos
    doctores = Doctor.objects.filter(activo=True).order_by('nombres')
    
    # Preparar datos
    doctores_data = []
    for doctor in doctores:
        consultas_doctor = consultas.filter(doctor=doctor)
        
        if not consultas_doctor.exists():
            continue
        
        # Construir nombre completo
        nombre_completo = f"{doctor.nombres} {doctor.apellidos}".strip()
        
        fila = {
            'doctor_nombre': nombre_completo,
            'total_consultas': consultas_doctor.count(),
            'total_ingresos': consultas_doctor.aggregate(total=Sum('precio'))['total'] or 0,
        }
        
        # Agrupar por tipo
        for tipo in tipos:
            consultas_tipo = consultas_doctor.filter(tipo_consulta=tipo)
            fila[f'tipo_{tipo}_cant'] = consultas_tipo.count()
            fila[f'tipo_{tipo}_total'] = consultas_tipo.aggregate(total=Sum('precio'))['total'] or 0
        
        doctores_data.append(fila)
    
    # Ordenar por total de ingresos
    doctores_data.sort(key=lambda x: x['total_ingresos'], reverse=True)
    
    # Paginación
    paginator = Paginator(doctores_data, 20)
    page_obj = paginator.get_page(page)
    
    return {
        'doctores': page_obj,

        'paginator': paginator,
        'page_obj': page_obj,
    }
######
def get_ingresos_consultas(desde, hasta, page=1):
    """Reporte de ingresos por consultas"""
    return {'mensaje': 'En desarrollo'}
################

def get_laboratorio(desde, hasta, page=1):
    """Reporte de laboratorio - Ingresos por tipo de examen"""
    from consultas.models import OrdenExamen
    from django.db.models import Sum, Count, Q
    from django.core.paginator import Paginator
    
    # Consulta base
    examenes = OrdenExamen.objects.filter(
        tipo_examen='LABORATORIO',
        fecha_solicitud__date__gte=desde,
        fecha_solicitud__date__lte=hasta,
        pagado=True  # Solo los pagados
    ).select_related('consulta__paciente', 'venta_ambulatoria')
    
    # Totales generales
    total_ingresos = examenes.aggregate(total=Sum('monto_pagado'))['total'] or 0
    total_examenes = examenes.count()
    
    # Separar internas vs ambulatorias
    internas = examenes.filter(consulta__isnull=False)
    ambulatorias = examenes.filter(consulta__isnull=True)
    
    total_internas = internas.aggregate(total=Sum('monto_pagado'))['total'] or 0
    total_ambulatorias = ambulatorias.aggregate(total=Sum('monto_pagado'))['total'] or 0
    
    # Ingresos por tipo de examen
    examenes_por_tipo = examenes.values('examen_especifico').annotate(
        cantidad=Count('id'),
        total_internas=Sum('monto_pagado', filter=Q(consulta__isnull=False)),
        total_ambulatorias=Sum('monto_pagado', filter=Q(consulta__isnull=True)),
        total_general=Sum('monto_pagado')
    ).order_by('-total_general')
    
    # Desglose por método de pago
    metodos_pago = examenes.values('metodo_pago').annotate(
        total_internas=Sum('monto_pagado', filter=Q(consulta__isnull=False)),
        total_ambulatorias=Sum('monto_pagado', filter=Q(consulta__isnull=True)),
        total_general=Sum('monto_pagado')
    ).order_by('-total_general')
    
    # Listado detallado con paginación
    paginator = Paginator(examenes.order_by('-fecha_solicitud'), 20)
    try:
        examenes_page = paginator.page(page)
    except:
        examenes_page = paginator.page(1)
    
    detalle = []
    for e in examenes_page:
        paciente = e.get_paciente_nombre()
        dni = e.get_paciente_dni()
        tipo = "Interno" if e.consulta else "Ambulatorio"
        
        detalle.append({
            'fecha': e.fecha_solicitud.strftime('%d/%m/%Y'),
            'hora': e.fecha_solicitud.strftime('%H:%M'),
            'paciente': paciente,
            'dni': dni,
            'examen': e.examen_especifico,
            'tipo': tipo,
            'metodo_pago': e.get_metodo_pago_display() if e.metodo_pago else '-',
            'monto': e.monto_pagado or 0,
        })
    
    return {
        'total_ingresos': total_ingresos,
        'total_examenes': total_examenes,
        'total_internas': total_internas,
        'total_ambulatorias': total_ambulatorias,
        'porcentaje_internas': (total_internas / total_ingresos * 100) if total_ingresos > 0 else 0,
        'porcentaje_ambulatorias': (total_ambulatorias / total_ingresos * 100) if total_ingresos > 0 else 0,
        'examenes_por_tipo': examenes_por_tipo,
        'metodos_pago': metodos_pago,
        'detalle': detalle,
        'total_registros': total_examenes,
        'paginator': paginator,
        'page_obj': examenes_page,
        'page': page,
    }
###############
def get_farmacia(desde, hasta, page=1):
    """Reporte de farmacia - SOLO medicamentos (con Top 20 Menos Vendidos)"""
    from farmacia.models import Venta, DetalleVenta, Medicamento, MovimientoInventario
    from django.db.models import Sum, Count, Q
    from django.core.paginator import Paginator
    from decimal import Decimal
    from collections import defaultdict
    
    print(f"\n===== REPORTE FARMACIA =====")
    print(f"Desde: {desde}, Hasta: {hasta}")
    
    # Ventas que tienen al menos un medicamento
    ventas_con_medicamentos = Venta.objects.filter(
        fecha_hora__date__gte=desde,
        fecha_hora__date__lte=hasta,
        detalles__medicamento__isnull=False
    ).select_related('usuario').prefetch_related('detalles__medicamento').distinct()
    
    print(f"Ventas con medicamentos encontradas: {ventas_con_medicamentos.count()}")
    
    # 🔥 CALCULAR TOTAL SOLO DE MEDICAMENTOS
    total_medicamentos = Decimal('0')
    for venta in ventas_con_medicamentos:
        suma_medicamentos = venta.detalles.filter(
            medicamento__isnull=False
        ).aggregate(total=Sum('subtotal'))['total'] or 0
        total_medicamentos += suma_medicamentos
    
    total_ventas = ventas_con_medicamentos.count()
    
    print(f"Total solo medicamentos: {total_medicamentos}")
    
    # 🔥 VENTAS POR VENDEDOR (AGRUPADAS CORRECTAMENTE)
    vendedores_dict = defaultdict(lambda: {'cantidad': 0, 'total': Decimal('0')})
    
    for venta in ventas_con_medicamentos:
        if venta.usuario:
            key = venta.usuario.id
            vendedores_dict[key]['username'] = venta.usuario.username
            vendedores_dict[key]['first_name'] = venta.usuario.first_name
            vendedores_dict[key]['last_name'] = venta.usuario.last_name
            vendedores_dict[key]['cantidad'] += 1
            
            suma_medicamentos = venta.detalles.filter(
                medicamento__isnull=False
            ).aggregate(total=Sum('subtotal'))['total'] or 0
            vendedores_dict[key]['total'] += suma_medicamentos
    
    ventas_por_vendedor = []
    for vendedor_id, data in vendedores_dict.items():
        ventas_por_vendedor.append({
            'usuario__username': data['username'],
            'usuario__first_name': data['first_name'],
            'usuario__last_name': data['last_name'],
            'cantidad': data['cantidad'],
            'total': float(data['total']),
        })
    
    # Ordenar por total descendente
    ventas_por_vendedor.sort(key=lambda x: x['total'], reverse=True)
    
    print(f"Vendedores únicos encontrados: {len(ventas_por_vendedor)}")
    for v in ventas_por_vendedor:
        print(f"  {v['usuario__first_name']}: {v['cantidad']} ventas - S/ {v['total']}")
    
    # ============================================
    # Top 20 MÁS VENDIDOS (ya existía)
    # ============================================
    ventas_por_medicamento = DetalleVenta.objects.filter(
        venta__in=ventas_con_medicamentos,
        medicamento__isnull=False
    ).values(
        'medicamento__nombre_comercial'
    ).annotate(
        cantidad=Sum('cantidad'),
        total=Sum('subtotal')
    ).order_by('-total')[:20]
    
    # ============================================
    # 🔥 NUEVO: Top 20 MENOS VENDIDOS
    # ============================================
    print(f"\n===== CALCULANDO TOP 20 MENOS VENDIDOS =====")
    menos_vendidos = []
    medicamentos = Medicamento.objects.filter(activo=True)
    
    for med in medicamentos:
        # Ventas en el período
        ventas = MovimientoInventario.objects.filter(
            medicamento=med,
            tipo='VENTA',
            fecha__date__gte=desde,
            fecha__date__lte=hasta
        ).aggregate(total=Sum('cantidad'))
        
        cantidad_vendida = ventas['total'] or 0
        
        # Última venta
        ultima_venta = MovimientoInventario.objects.filter(
            medicamento=med,
            tipo='VENTA'
        ).order_by('-fecha').first()
        
        dias_sin_venta = 999
        if ultima_venta:
            dias_sin_venta = (hasta - ultima_venta.fecha.date()).days
            if dias_sin_venta < 0:
                dias_sin_venta = 0
        
        # Rotación (ventas / stock actual * 100)
        rotacion = 0
        if med.stock_actual > 0:
            rotacion = (cantidad_vendida / med.stock_actual) * 100
        
        # Valor del stock inmovilizado
        valor_stock = float(med.stock_actual * med.precio_compra)
        
        menos_vendidos.append({
            'id': med.id,
            'codigo': med.codigo,
            'nombre': med.nombre_comercial,
            'stock_actual': med.stock_actual,
            'cantidad_vendida': cantidad_vendida,
            'dias_sin_venta': dias_sin_venta,
            'rotacion': round(rotacion, 2),
            'valor_stock': valor_stock,
        })
    
    # Ordenar por cantidad_vendida (ascendente = menos vendidos primero)
    # Si tienen la misma cantidad, ordenar por mayor valor de stock
    menos_vendidos.sort(key=lambda x: (x['cantidad_vendida'], -x['valor_stock']))
    top_20_menos_vendidos = menos_vendidos[:20]
    
    print(f"Total medicamentos activos: {len(medicamentos)}")
    print(f"Medicamentos sin ventas: {len([m for m in menos_vendidos if m['cantidad_vendida'] == 0])}")
    print(f"Top 20 menos vendidos calculados")
    
    # ============================================
    # Desglose por método de pago
    # ============================================
    metodos_pago = []
    for metodo in ['EFECTIVO', 'YAPE', 'PLIN']:
        ventas_metodo = ventas_con_medicamentos.filter(metodo_pago=metodo)
        total_metodo = Decimal('0')
        for venta in ventas_metodo:
            total_metodo += venta.detalles.filter(
                medicamento__isnull=False
            ).aggregate(total=Sum('subtotal'))['total'] or 0
        
        if total_metodo > 0:
            metodos_pago.append({
                'metodo_pago': metodo,
                'total': float(total_metodo),
            })
    
    # ============================================
    # Listado detallado con paginación
    # ============================================
    paginator = Paginator(ventas_con_medicamentos.order_by('-fecha_hora'), 20)
    try:
        ventas_page = paginator.page(page)
    except:
        ventas_page = paginator.page(1)
    
    detalle = []
    for v in ventas_page:
        vendedor = v.usuario.get_full_name() or v.usuario.username if v.usuario else "Sistema"
        
        items = v.detalles.filter(medicamento__isnull=False)
        
        if items.exists():
            medicamentos_list = []
            total_venta_medicamentos = Decimal('0')
            
            for d in items:
                medicamentos_list.append(f"{d.medicamento.nombre_comercial} x{d.cantidad}")
                total_venta_medicamentos += d.subtotal
            
            medicamentos = ", ".join(medicamentos_list[:3])
            if items.count() > 3:
                medicamentos += f" y {items.count()-3} más"
            
            detalle.append({
                'fecha': v.fecha_hora.strftime('%d/%m/%Y'),
                'hora': v.fecha_hora.strftime('%H:%M'),
                'vendedor': vendedor,
                'metodo_pago': v.get_metodo_pago_display(),
                'medicamentos': medicamentos,
                'total': float(total_venta_medicamentos),
            })
    
    # ============================================
    # Retornar todos los datos
    # ============================================
    return {
        'total_ingresos': float(total_medicamentos),
        'total_ventas': total_ventas,
        'ventas_por_vendedor': ventas_por_vendedor,
        'ventas_por_medicamento': ventas_por_medicamento,
        'menos_vendidos': top_20_menos_vendidos,  # 🔥 NUEVO CAMPO
        'metodos_pago': metodos_pago,
        'detalle': detalle,
        'total_registros': total_ventas,
        'paginator': paginator,
        'page_obj': ventas_page,
        'page': page,
    }

##############
def get_topico(desde, hasta, page=1):
    """Reporte de tópico - Mano de obra y procedimientos"""
    from farmacia.models import Venta, DetalleVenta
    from django.db.models import Sum, Count, Q
    from django.core.paginator import Paginator
    from decimal import Decimal

    print(f"\n===== REPORTE TÓPICO =====")
    print(f"Desde: {desde}, Hasta: {hasta}")

    # Ventas que tienen al menos un detalle SIN medicamento (tópico)
    ventas = Venta.objects.filter(
        fecha_hora__date__gte=desde,
        fecha_hora__date__lte=hasta,
        detalles__medicamento__isnull=True
    ).select_related('usuario').prefetch_related('detalles').distinct()

    print(f"Ventas encontradas: {ventas.count()}")

    # Calcular total de ingresos SUMANDO LOS SUBTOTALES de detalles sin medicamento
    total_ingresos = Decimal('0')
    for v in ventas:
        suma_topico = v.detalles.filter(
            medicamento__isnull=True
        ).aggregate(total=Sum('subtotal'))['total'] or 0
        total_ingresos += suma_topico
        print(f"  Venta #{v.id}: Total venta: {v.total}, Tópico: {suma_topico}")

    total_procedimientos = DetalleVenta.objects.filter(
        venta__in=ventas,
        medicamento__isnull=True
    ).count()

    print(f"Total ingresos calculado: {total_ingresos}")
    print(f"Total procedimientos: {total_procedimientos}")

    # Procedimientos por tipo
    procedimientos_por_tipo = DetalleVenta.objects.filter(
        venta__in=ventas,
        medicamento__isnull=True
    ).values('descripcion').annotate(
        cantidad=Count('id'),
        total=Sum('subtotal')
    ).order_by('-total')

    # Ventas por vendedor
    ventas_por_vendedor = ventas.values(
        'usuario__username', 'usuario__first_name', 'usuario__last_name'
    ).annotate(
        cantidad=Count('id'),
        total=Sum('total')
    ).order_by('-total')

    # Desglose por método de pago
    metodos_pago = ventas.values('metodo_pago').annotate(
        total=Sum('total')
    ).order_by('-total')

    # Listado detallado con paginación
    paginator = Paginator(ventas.order_by('-fecha_hora'), 20)
    try:
        ventas_page = paginator.page(page)
    except:
        ventas_page = paginator.page(1)

    detalle = []
    for v in ventas_page:
        vendedor = v.usuario.get_full_name() or v.usuario.username if v.usuario else "Sistema"
        items = v.detalles.filter(medicamento__isnull=True)

        procedimientos_list = []
        total_venta = 0
        for d in items:
            nombre = d.descripcion or "Procedimiento"
            procedimientos_list.append(f"{nombre} x{d.cantidad}")
            total_venta += d.subtotal

        procedimientos = ", ".join(procedimientos_list[:3])
        if items.count() > 3:
            procedimientos += f" y {items.count()-3} más"

        detalle.append({
            'fecha': v.fecha_hora.strftime('%d/%m/%Y'),
            'hora': v.fecha_hora.strftime('%H:%M'),
            'vendedor': vendedor,
            'metodo_pago': v.get_metodo_pago_display(),
            'procedimientos': procedimientos,
            'total': float(total_venta),
        })

    return {
        'total_ingresos': float(total_ingresos),
        'total_procedimientos': total_procedimientos,
        'procedimientos_por_tipo': procedimientos_por_tipo,
        'ventas_por_vendedor': ventas_por_vendedor,
        'metodos_pago': metodos_pago,
        'detalle': detalle,
        'total_registros': ventas.count(),
        'paginator': paginator,
        'page_obj': ventas_page,
        'page': page,
    }

###########################
def get_topico(desde, hasta, page=1):
    """Reporte de tópico - Mano de obra y procedimientos"""
    from farmacia.models import Venta, DetalleVenta
    from django.db.models import Sum, Count, Q
    from django.core.paginator import Paginator
    from decimal import Decimal
    from collections import defaultdict

    print(f"\n===== REPORTE TÓPICO =====")
    print(f"Desde: {desde}, Hasta: {hasta}")

    # Ventas que tienen al menos un detalle SIN medicamento (tópico)
    ventas = Venta.objects.filter(
        fecha_hora__date__gte=desde,
        fecha_hora__date__lte=hasta,
        detalles__medicamento__isnull=True
    ).select_related('usuario').prefetch_related('detalles').distinct()

    print(f"Ventas encontradas: {ventas.count()}")

    # Calcular total de ingresos (SOLO tópico)
    total_ingresos = Decimal('0')
    for v in ventas:
        suma_topico = v.detalles.filter(
            medicamento__isnull=True
        ).aggregate(total=Sum('subtotal'))['total'] or 0
        total_ingresos += suma_topico
        print(f"  Venta #{v.id}: Tópico: {suma_topico}")

    total_procedimientos = DetalleVenta.objects.filter(
        venta__in=ventas,
        medicamento__isnull=True
    ).count()

    print(f"Total ingresos calculado: {total_ingresos}")
    print(f"Total procedimientos: {total_procedimientos}")

    # Procedimientos por tipo
    procedimientos_por_tipo = DetalleVenta.objects.filter(
        venta__in=ventas,
        medicamento__isnull=True
    ).values('descripcion').annotate(
        cantidad=Count('id'),
        total=Sum('subtotal')
    ).order_by('-total')

    # 🔥 VENTAS POR VENDEDOR (AGRUPADAS CORRECTAMENTE)
    vendedores_dict = defaultdict(lambda: {'cantidad': 0, 'total': 0})

    for venta in ventas:
        if venta.usuario:
            usuario = venta.usuario
            key = usuario.id
            
            # Calcular el tópico de esta venta
            suma_topico = venta.detalles.filter(
                medicamento__isnull=True
            ).aggregate(total=Sum('subtotal'))['total'] or 0
            
            vendedores_dict[key]['username'] = usuario.username
            vendedores_dict[key]['first_name'] = usuario.first_name
            vendedores_dict[key]['last_name'] = usuario.last_name
            vendedores_dict[key]['cantidad'] += 1
            vendedores_dict[key]['total'] += float(suma_topico)

    # Convertir a lista para el template
    ventas_por_vendedor = []
    for vendedor_id, data in vendedores_dict.items():
        ventas_por_vendedor.append({
            'usuario__username': data['username'],
            'usuario__first_name': data['first_name'],
            'usuario__last_name': data['last_name'],
            'cantidad': data['cantidad'],
            'total': data['total'],
        })

    # Ordenar por total descendente
    ventas_por_vendedor.sort(key=lambda x: x['total'], reverse=True)

    print(f"Vendedores únicos encontrados: {len(ventas_por_vendedor)}")
    for v in ventas_por_vendedor:
        print(f"  {v['usuario__first_name']}: {v['cantidad']} ventas - S/ {v['total']}")

    # Desglose por método de pago (SOLO tópico)
    metodos_pago = []
    for metodo in ['EFECTIVO', 'YAPE', 'PLIN']:
        ventas_metodo = ventas.filter(metodo_pago=metodo)
        total_metodo = Decimal('0')
        for venta in ventas_metodo:
            total_metodo += venta.detalles.filter(
                medicamento__isnull=True
            ).aggregate(total=Sum('subtotal'))['total'] or 0
        
        if total_metodo > 0:
            metodos_pago.append({
                'metodo_pago': metodo,
                'total': float(total_metodo),
            })

    # Listado detallado con paginación
    paginator = Paginator(ventas.order_by('-fecha_hora'), 20)
    try:
        ventas_page = paginator.page(page)
    except:
        ventas_page = paginator.page(1)

    detalle = []
    for v in ventas_page:
        vendedor = v.usuario.get_full_name() or v.usuario.username if v.usuario else "Sistema"
        items = v.detalles.filter(medicamento__isnull=True)

        procedimientos_list = []
        total_venta = 0
        for d in items:
            nombre = d.descripcion or "Procedimiento"
            procedimientos_list.append(f"{nombre} x{d.cantidad}")
            total_venta += d.subtotal

        procedimientos = ", ".join(procedimientos_list[:3])
        if items.count() > 3:
            procedimientos += f" y {items.count()-3} más"

        detalle.append({
            'fecha': v.fecha_hora.strftime('%d/%m/%Y'),
            'hora': v.fecha_hora.strftime('%H:%M'),
            'vendedor': vendedor,
            'metodo_pago': v.get_metodo_pago_display(),
            'procedimientos': procedimientos,
            'total': float(total_venta),
        })

    return {
        'total_ingresos': float(total_ingresos),
        'total_procedimientos': total_procedimientos,
        'procedimientos_por_tipo': procedimientos_por_tipo,
        'ventas_por_vendedor': ventas_por_vendedor,
        'metodos_pago': metodos_pago,
        'detalle': detalle,
        'total_registros': ventas.count(),
        'paginator': paginator,
        'page_obj': ventas_page,
        'page': page,
    }

########
def get_ingresos_totales(desde, hasta, page=1):
    """Reporte de ingresos totales - Suma de todos los módulos (INCLUYE TÓPICO)"""
    from consultas.models import Consulta, OrdenExamen
    from farmacia.models import Venta, DetalleVenta
    from pagos.models import OrdenPago
    from django.db.models import Sum, Q
    from decimal import Decimal
    
    print(f"\n===== INGRESOS TOTALES =====")
    print(f"Desde: {desde}, Hasta: {hasta}")
    
    # 1. INGRESOS POR CONSULTAS
    pagos_consultas = OrdenPago.objects.filter(
        consulta__isnull=False,
        fecha_pago__date__gte=desde,
        fecha_pago__date__lte=hasta,
        estado='PAGADO'
    )
    total_consultas = pagos_consultas.aggregate(total=Sum('monto'))['total'] or 0
    cantidad_consultas = pagos_consultas.count()
    
    print(f"Consultas: {cantidad_consultas} pagos - S/ {total_consultas}")
    
    # 2. INGRESOS POR LABORATORIO
    pagos_lab = OrdenExamen.objects.filter(
        pagado=True,
        fecha_pago__date__gte=desde,
        fecha_pago__date__lte=hasta
    )
    total_laboratorio = pagos_lab.aggregate(total=Sum('monto_pagado'))['total'] or 0
    cantidad_lab = pagos_lab.count()
    
    print(f"Laboratorio: {cantidad_lab} exámenes - S/ {total_laboratorio}")
    
    # 3. INGRESOS POR FARMACIA (SOLO medicamentos)
    # Calculamos sumando los subtotales de detalles con medicamento
    total_farmacia = Decimal('0')
    ventas_farmacia = Venta.objects.filter(
        fecha_hora__date__gte=desde,
        fecha_hora__date__lte=hasta
    )
    
    for venta in ventas_farmacia:
        suma_medicamentos = venta.detalles.filter(
            medicamento__isnull=False
        ).aggregate(total=Sum('subtotal'))['total'] or 0
        total_farmacia += suma_medicamentos
    
    cantidad_farmacia = ventas_farmacia.filter(
        detalles__medicamento__isnull=False
    ).distinct().count()
    
    print(f"Farmacia: {cantidad_farmacia} ventas - S/ {total_farmacia}")
    
    # 4. INGRESOS POR TÓPICO (SOLO mano de obra)
    total_topico = Decimal('0')
    for venta in ventas_farmacia:
        suma_topico = venta.detalles.filter(
            medicamento__isnull=True
        ).aggregate(total=Sum('subtotal'))['total'] or 0
        total_topico += suma_topico
    
    cantidad_topico = ventas_farmacia.filter(
        detalles__medicamento__isnull=True
    ).distinct().count()
    
    print(f"Tópico: {cantidad_topico} procedimientos - S/ {total_topico}")
    
    # 5. TOTAL GENERAL
    total_general = total_consultas + total_laboratorio + total_farmacia + total_topico
    
    # 6. INGRESOS POR MÉTODO DE PAGO (global)
    # Consultas
    efectivo_consultas = pagos_consultas.filter(metodo_pago='EFECTIVO').aggregate(total=Sum('monto'))['total'] or 0
    yape_consultas = pagos_consultas.filter(metodo_pago='YAPE').aggregate(total=Sum('monto'))['total'] or 0
    plin_consultas = pagos_consultas.filter(metodo_pago='PLIN').aggregate(total=Sum('monto'))['total'] or 0
    
    # Laboratorio
    efectivo_lab = pagos_lab.filter(metodo_pago='EFECTIVO').aggregate(total=Sum('monto_pagado'))['total'] or 0
    yape_lab = pagos_lab.filter(metodo_pago='YAPE').aggregate(total=Sum('monto_pagado'))['total'] or 0
    plin_lab = pagos_lab.filter(metodo_pago='PLIN').aggregate(total=Sum('monto_pagado'))['total'] or 0
    
    # Farmacia (solo medicamentos)
    efectivo_farmacia = Decimal('0')
    yape_farmacia = Decimal('0')
    plin_farmacia = Decimal('0')
    
    for venta in ventas_farmacia.filter(metodo_pago='EFECTIVO'):
        efectivo_farmacia += venta.detalles.filter(medicamento__isnull=False).aggregate(total=Sum('subtotal'))['total'] or 0
    
    for venta in ventas_farmacia.filter(metodo_pago='YAPE'):
        yape_farmacia += venta.detalles.filter(medicamento__isnull=False).aggregate(total=Sum('subtotal'))['total'] or 0
    
    for venta in ventas_farmacia.filter(metodo_pago='PLIN'):
        plin_farmacia += venta.detalles.filter(medicamento__isnull=False).aggregate(total=Sum('subtotal'))['total'] or 0
    
    # Tópico
    efectivo_topico = Decimal('0')
    yape_topico = Decimal('0')
    plin_topico = Decimal('0')
    
    for venta in ventas_farmacia.filter(metodo_pago='EFECTIVO'):
        efectivo_topico += venta.detalles.filter(medicamento__isnull=True).aggregate(total=Sum('subtotal'))['total'] or 0
    
    for venta in ventas_farmacia.filter(metodo_pago='YAPE'):
        yape_topico += venta.detalles.filter(medicamento__isnull=True).aggregate(total=Sum('subtotal'))['total'] or 0
    
    for venta in ventas_farmacia.filter(metodo_pago='PLIN'):
        plin_topico += venta.detalles.filter(medicamento__isnull=True).aggregate(total=Sum('subtotal'))['total'] or 0
    
    # Totales por método
    total_efectivo = efectivo_consultas + efectivo_lab + efectivo_farmacia + efectivo_topico
    total_yape = yape_consultas + yape_lab + yape_farmacia + yape_topico
    total_plin = plin_consultas + plin_lab + plin_farmacia + plin_topico
    
    # Calcular porcentajes
    porcentaje_consultas = (total_consultas / total_general * 100) if total_general > 0 else 0
    porcentaje_laboratorio = (total_laboratorio / total_general * 100) if total_general > 0 else 0
    porcentaje_farmacia = (total_farmacia / total_general * 100) if total_general > 0 else 0
    porcentaje_topico = (total_topico / total_general * 100) if total_general > 0 else 0
    
    print(f"\nRESUMEN FINAL:")
    print(f"Consultas: S/ {total_consultas}")
    print(f"Laboratorio: S/ {total_laboratorio}")
    print(f"Farmacia: S/ {total_farmacia}")
    print(f"Tópico: S/ {total_topico}")
    print(f"TOTAL GENERAL: S/ {total_general}")
    
    return {
        'total_consultas': float(total_consultas),
        'cantidad_consultas': cantidad_consultas,
        'total_laboratorio': float(total_laboratorio),
        'cantidad_lab': cantidad_lab,
        'total_farmacia': float(total_farmacia),
        'cantidad_farmacia': cantidad_farmacia,
        'total_topico': float(total_topico),
        'cantidad_topico': cantidad_topico,
        'total_general': float(total_general),
        'total_efectivo': float(total_efectivo),
        'total_yape': float(total_yape),
        'total_plin': float(total_plin),
        'porcentaje_consultas': porcentaje_consultas,
        'porcentaje_laboratorio': porcentaje_laboratorio,
        'porcentaje_farmacia': porcentaje_farmacia,
        'porcentaje_topico': porcentaje_topico,
    }

###################
def get_ingresos_totales(desde, hasta, page=1):
    """Reporte de ingresos totales - Suma de todos los módulos (INCLUYE TÓPICO) SIN PORCENTAJES"""
    from consultas.models import Consulta, OrdenExamen
    from farmacia.models import Venta, DetalleVenta
    from pagos.models import OrdenPago
    from django.db.models import Sum, Q
    from decimal import Decimal
    
    print(f"\n===== INGRESOS TOTALES =====")
    print(f"Desde: {desde}, Hasta: {hasta}")
    
    # 1. INGRESOS POR CONSULTAS
    pagos_consultas = OrdenPago.objects.filter(
        consulta__isnull=False,
        fecha_pago__date__gte=desde,
        fecha_pago__date__lte=hasta,
        estado='PAGADO'
    )
    total_consultas = pagos_consultas.aggregate(total=Sum('monto'))['total'] or 0
    cantidad_consultas = pagos_consultas.count()
    
    print(f"Consultas: {cantidad_consultas} pagos - S/ {total_consultas}")
    
    # 2. INGRESOS POR LABORATORIO
    pagos_lab = OrdenExamen.objects.filter(
        pagado=True,
        fecha_pago__date__gte=desde,
        fecha_pago__date__lte=hasta
    )
    total_laboratorio = pagos_lab.aggregate(total=Sum('monto_pagado'))['total'] or 0
    cantidad_lab = pagos_lab.count()
    
    print(f"Laboratorio: {cantidad_lab} exámenes - S/ {total_laboratorio}")
    
    # 3. INGRESOS POR FARMACIA (SOLO medicamentos)
    ventas_farmacia = Venta.objects.filter(
        fecha_hora__date__gte=desde,
        fecha_hora__date__lte=hasta
    )
    
    total_farmacia = Decimal('0')
    for venta in ventas_farmacia:
        suma_medicamentos = venta.detalles.filter(
            medicamento__isnull=False
        ).aggregate(total=Sum('subtotal'))['total'] or 0
        total_farmacia += suma_medicamentos
    
    cantidad_farmacia = ventas_farmacia.filter(
        detalles__medicamento__isnull=False
    ).distinct().count()
    
    print(f"Farmacia: {cantidad_farmacia} ventas - S/ {total_farmacia}")
    
    # 4. INGRESOS POR TÓPICO (SOLO mano de obra)
    total_topico = Decimal('0')
    ventas_con_topico = ventas_farmacia.filter(detalles__medicamento__isnull=True).distinct()
    
    print(f"Ventas con tópico encontradas: {ventas_con_topico.count()}")
    
    for venta in ventas_con_topico:
        suma_topico = venta.detalles.filter(
            medicamento__isnull=True
        ).aggregate(total=Sum('subtotal'))['total'] or 0
        total_topico += suma_topico
        print(f"  Venta #{venta.id}: S/ {suma_topico}")
    
    cantidad_topico = ventas_con_topico.count()
    print(f"Total tópico calculado: S/ {total_topico}")
    
    # 5. TOTAL GENERAL
    total_general = total_consultas + total_laboratorio + total_farmacia + total_topico
    
    return {
        'total_consultas': float(total_consultas),
        'cantidad_consultas': cantidad_consultas,
        'total_laboratorio': float(total_laboratorio),
        'cantidad_lab': cantidad_lab,
        'total_farmacia': float(total_farmacia),
        'cantidad_farmacia': cantidad_farmacia,
        'total_topico': float(total_topico),
        'cantidad_topico': cantidad_topico,
        'total_general': float(total_general),
    }


#######################
def get_metodos_pago(desde, hasta, page=1):
    """Reporte de métodos de pago - Resumen por método (CON TÓPICO SEPARADO)"""
    from consultas.models import Consulta, OrdenExamen
    from farmacia.models import Venta, DetalleVenta
    from pagos.models import OrdenPago
    from django.db.models import Sum
    from decimal import Decimal

    print(f"\n===== MÉTODOS DE PAGO =====")
    print(f"Desde: {desde}, Hasta: {hasta}")

    # 1. CONSULTAS
    pagos_consultas = OrdenPago.objects.filter(
        consulta__isnull=False,
        fecha_pago__date__gte=desde,
        fecha_pago__date__lte=hasta,
        estado='PAGADO'
    )

    efectivo_cons = pagos_consultas.filter(metodo_pago='EFECTIVO').aggregate(total=Sum('monto'))['total'] or 0
    yape_cons = pagos_consultas.filter(metodo_pago='YAPE').aggregate(total=Sum('monto'))['total'] or 0
    plin_cons = pagos_consultas.filter(metodo_pago='PLIN').aggregate(total=Sum('monto'))['total'] or 0

    # 2. LABORATORIO
    pagos_lab = OrdenExamen.objects.filter(
        pagado=True,
        fecha_pago__date__gte=desde,
        fecha_pago__date__lte=hasta
    )

    efectivo_lab = pagos_lab.filter(metodo_pago='EFECTIVO').aggregate(total=Sum('monto_pagado'))['total'] or 0
    yape_lab = pagos_lab.filter(metodo_pago='YAPE').aggregate(total=Sum('monto_pagado'))['total'] or 0
    plin_lab = pagos_lab.filter(metodo_pago='PLIN').aggregate(total=Sum('monto_pagado'))['total'] or 0

    # 3. FARMACIA (SOLO medicamentos)
    ventas = Venta.objects.filter(
        fecha_hora__date__gte=desde,
        fecha_hora__date__lte=hasta
    ).prefetch_related('detalles')

    efectivo_far = Decimal('0')
    yape_far = Decimal('0')
    plin_far = Decimal('0')
    
    efectivo_top = Decimal('0')
    yape_top = Decimal('0')
    plin_top = Decimal('0')

    for venta in ventas:
        # Separar medicamentos vs tópico
        suma_medicamentos = venta.detalles.filter(
            medicamento__isnull=False
        ).aggregate(total=Sum('subtotal'))['total'] or 0
        
        suma_topico = venta.detalles.filter(
            medicamento__isnull=True
        ).aggregate(total=Sum('subtotal'))['total'] or 0
        
        if venta.metodo_pago == 'EFECTIVO':
            efectivo_far += suma_medicamentos
            efectivo_top += suma_topico
        elif venta.metodo_pago == 'YAPE':
            yape_far += suma_medicamentos
            yape_top += suma_topico
        elif venta.metodo_pago == 'PLIN':
            plin_far += suma_medicamentos
            plin_top += suma_topico

    # TOTALES POR MÉTODO
    total_efectivo = efectivo_cons + efectivo_lab + efectivo_far + efectivo_top
    total_yape = yape_cons + yape_lab + yape_far + yape_top
    total_plin = plin_cons + plin_lab + plin_far + plin_top
    total_general = total_efectivo + total_yape + total_plin

    # Datos para el template (CON TÓPICO SEPARADO)
    metodos = [
        {
            'metodo': 'Efectivo', 
            'total': float(total_efectivo),
            'consultas': float(efectivo_cons), 
            'laboratorio': float(efectivo_lab), 
            'farmacia': float(efectivo_far),
            'topico': float(efectivo_top),
        },
        {
            'metodo': 'Yape', 
            'total': float(total_yape),
            'consultas': float(yape_cons), 
            'laboratorio': float(yape_lab), 
            'farmacia': float(yape_far),
            'topico': float(yape_top),
        },
        {
            'metodo': 'Plin', 
            'total': float(total_plin),
            'consultas': float(plin_cons), 
            'laboratorio': float(plin_lab), 
            'farmacia': float(plin_far),
            'topico': float(plin_top),
        },
    ]

    return {
        'metodos': metodos,
        'total_efectivo': float(total_efectivo),
        'total_yape': float(total_yape),
        'total_plin': float(total_plin),
        'total_general': float(total_general),
        }    


##################
def get_inventario(request=None, page=1):
    """Reporte de inventario - Stock bajo y próximos a vencer (4 meses)"""
    from farmacia.models import Medicamento
    from django.db.models import F, Sum, Q
    from django.core.paginator import Paginator
    from datetime import date, timedelta
    from decimal import Decimal

    print(f"\n===== INVENTARIO =====")
    
    # Obtener término de búsqueda si existe
    query = ''
    if request and hasattr(request, 'GET') and request.GET.get('q'):
        query = request.GET.get('q')
        print(f"Buscando: '{query}'")
    
    # Base de medicamentos activos
    medicamentos_base = Medicamento.objects.filter(activo=True)
    
    # Aplicar búsqueda si existe
    if query:
        medicamentos_base = medicamentos_base.filter(
            Q(nombre_comercial__icontains=query) |
            Q(codigo__icontains=query) |
            Q(principio_activo__icontains=query)
        )

    # Stock bajo (stock actual < stock mínimo)
    stock_bajo = medicamentos_base.filter(
        stock_actual__lt=F('stock_minimo')
    ).order_by('stock_actual')

    # ============================================
    # 🔥 PRÓXIMOS A VENCER - AHORA 4 MESES (120 DÍAS)
    # ============================================
    hoy = date.today()
    limite_4_meses = hoy + timedelta(days=120)  # 4 MESES EXACTOS

    proximos_vencer = medicamentos_base.filter(
        fecha_vencimiento__lte=limite_4_meses,  # Cambiado de 30 a 120 días
        fecha_vencimiento__gte=hoy
    ).order_by('fecha_vencimiento')

    # Totales generales
    total_unidades = Medicamento.objects.filter(activo=True).aggregate(total=Sum('stock_actual'))['total'] or 0
    total_productos = Medicamento.objects.filter(activo=True).count()
    
    # Calcular valor total del inventario
    valor_total = Decimal('0')
    for m in Medicamento.objects.filter(activo=True):
        valor_total += m.stock_actual * m.precio_compra

    # Paginación para stock bajo
    paginator_stock = Paginator(stock_bajo, 20)
    try:
        stock_page = paginator_stock.page(page)
    except:
        stock_page = paginator_stock.page(1)

    # Paginación para próximos a vencer
    paginator_vencer = Paginator(proximos_vencer, 20)
    try:
        vencer_page = paginator_vencer.page(page)
    except:
        vencer_page = paginator_vencer.page(1)

    # ============================================
    # PREPARAR DATOS CON CLASIFICACIÓN POR URGENCIA
    # ============================================
    
    # Datos para stock bajo (sin cambios)
    stock_bajo_data = []
    for m in stock_page:
        stock_bajo_data.append({
            'id': m.id,
            'codigo': m.codigo,
            'nombre': m.nombre_comercial,
            'principio_activo': m.principio_activo,
            'stock_actual': m.stock_actual,
            'stock_minimo': m.stock_minimo,
            'diferencia': m.stock_minimo - m.stock_actual,
            'estado': 'CRÍTICO' if m.stock_actual == 0 else 'BAJO',
            'url_ajustar': f"/farmacia/medicamentos/ajustar/{m.id}/",
        })

    # Datos para próximos a vencer (CON CLASIFICACIÓN MEJORADA)
    proximos_vencer_data = []
    for m in vencer_page:
        dias = (m.fecha_vencimiento - hoy).days
        
        # Clasificar por urgencia
        if dias <= 30:
            urgencia = 'URGENTE'
            bg_color = 'danger'
        elif dias <= 60:
            urgencia = 'PRÓXIMO'
            bg_color = 'warning'
        else:
            urgencia = 'PRECAUCIÓN'
            bg_color = 'info'
        
        proximos_vencer_data.append({
            'id': m.id,
            'codigo': m.codigo,
            'nombre': m.nombre_comercial,
            'lote': m.lote or '-',
            'fecha_vencimiento': m.fecha_vencimiento.strftime('%d/%m/%Y'),
            'dias': dias,
            'urgencia': urgencia,
            'bg_color': bg_color,
            'url_detalle': f"/farmacia/medicamentos/detalle/{m.id}/",
            'url_ajustar': f"/farmacia/medicamentos/ajustar/{m.id}/",
        })

    print(f"Total productos: {total_productos}")
    print(f"Stock bajo: {stock_bajo.count()}")
    print(f"Próximos a vencer (4 meses): {proximos_vencer.count()}")
    print(f"  - Urgentes (<30d): {len([p for p in proximos_vencer_data if p['dias'] <= 30])}")
    print(f"  - Próximos (30-60d): {len([p for p in proximos_vencer_data if 30 < p['dias'] <= 60])}")
    print(f"  - Precaución (60-120d): {len([p for p in proximos_vencer_data if 60 < p['dias'] <= 120])}")

    return {
        'query': query,
        'stock_bajo': stock_bajo_data,
        'proximos_vencer': proximos_vencer_data,
        'total_stock_bajo': stock_bajo.count(),
        'total_proximos_vencer': proximos_vencer.count(),
        'limite_4_meses': limite_4_meses.strftime('%d/%m/%Y'),  # NUEVO
        'total_unidades': total_unidades,
        'total_productos': total_productos,
        'valor_total': float(valor_total),
        'paginator_stock': paginator_stock,
        'page_obj_stock': stock_page,
        'paginator_vencer': paginator_vencer,
        'page_obj_vencer': vencer_page,
        'page': page,
    }


################# excel #############



######################################
# ============================================
# FUNCIONES DE EXPORTACIÓN A EXCEL - COMPLETAS
# ============================================
#############################
def exportar_consultas_periodo(ws, data, header_font, header_fill, border):
    """Exportar consultas por período a Excel"""
    row = 3
    
    # Totales
    ws[f'A{row}'] = "TOTAL CONSULTAS:"
    ws[f'B{row}'] = data.get('total_consultas', 0)
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "ATENDIDAS:"
    ws[f'B{row}'] = data.get('atendidas', 0)
    
    row += 1
    ws[f'A{row}'] = "INGRESOS TOTALES:"
    total_ingresos = data.get('total_ingresos', 0)
    ws[f'B{row}'] = float(total_ingresos) if total_ingresos is not None else 0
    ws[f'B{row}'].number_format = '"S/ "#,##0.00'
    
    # Métodos de pago
    row += 2
    ws[f'A{row}'] = "INGRESOS POR MÉTODO DE PAGO"
    ws[f'A{row}'].font = Font(bold=True, italic=True)
    
    row += 1
    headers = ['Método', 'Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    metodos = [
        ('Efectivo', data.get('efectivo', 0)),
        ('Yape', data.get('yape', 0)),
        ('Plin', data.get('plin', 0)),
    ]
    
    for m in metodos:
        row += 1
        ws.cell(row=row, column=1).value = m[0]
        valor = m[1]
        ws.cell(row=row, column=2).value = float(valor) if valor is not None else 0
        ws.cell(row=row, column=2).number_format = '"S/ "#,##0.00'
        for col in [1, 2]:
            ws.cell(row=row, column=col).border = border
    
    # Listado detallado
    row += 2
    ws[f'A{row}'] = "LISTADO DETALLADO DE CONSULTAS"
    ws[f'A{row}'].font = Font(bold=True, italic=True)
    
    row += 1
    headers = ['Fecha', 'Hora', 'Paciente', 'Doctor', 'Tipo', 'Estado']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    for item in data.get('detalle', []):
        row += 1
        ws.cell(row=row, column=1).value = item.get('fecha', '')
        ws.cell(row=row, column=2).value = item.get('hora', '')
        ws.cell(row=row, column=3).value = item.get('paciente', '')
        ws.cell(row=row, column=4).value = item.get('doctor', '')
        ws.cell(row=row, column=5).value = item.get('tipo', '')
        ws.cell(row=row, column=6).value = item.get('estado', '')
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border
    
    # Ajustar ancho
    for col in range(1, 7):
        ws.column_dimensions[chr(64+col)].width = 20
###############################
def exportar_consultas_doctor(ws, data, header_font, header_fill, border):
    """Exportar consultas por doctor a Excel"""
    row = 3
    
    # Total de consultas (si viene en los datos)
    ws[f'A{row}'] = "TOTAL CONSULTAS EN EL PERÍODO:"
    # Calcular total de consultas sumando
    total_consultas = sum(doctor.get('total_consultas', 0) for doctor in data.get('doctores', []))
    ws[f'B{row}'] = total_consultas
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 2
    ws[f'A{row}'] = "CONSULTAS POR MÉDICO"
    ws[f'A{row}'].font = Font(bold=True, italic=True)
    
    row += 1
    headers = ['Médico', 'Cantidad Consultas', 'Total Ingresos (S/)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        if header_font:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Usar los datos de la página actual
    doctores_data = data.get('doctores', [])
    if hasattr(doctores_data, 'object_list'):  # Si es un Page
        doctores_list = doctores_data.object_list
    else:
        doctores_list = doctores_data
    
    for doctor in doctores_list:
        row += 1
        # Nombre ya viene como doctor_nombre
        ws.cell(row=row, column=1).value = doctor.get('doctor_nombre', '')
        ws.cell(row=row, column=2).value = doctor.get('total_consultas', 0)
        ws.cell(row=row, column=3).value = float(doctor.get('total_ingresos', 0))
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        for col in range(1, 4):
            if border:
                ws.cell(row=row, column=col).border = border
    
    # Totales
    row += 2
    ws[f'A{row}'] = "TOTALES:"
    ws[f'A{row}'].font = Font(bold=True)
    total_cantidad = sum(d.get('total_consultas', 0) for d in doctores_list)
    total_ingresos = sum(d.get('total_ingresos', 0) for d in doctores_list)
    ws[f'B{row}'] = total_cantidad
    ws[f'C{row}'] = float(total_ingresos)
    ws[f'C{row}'].number_format = '#,##0.00'
    ws[f'A{row}'].border = border
    ws[f'B{row}'].border = border
    ws[f'C{row}'].border = border
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
#########################

def exportar_laboratorio(ws, data, header_font, header_fill, border):
    """Exportar laboratorio a Excel - CON VALIDACIÓN DE NULOS"""
    row = 3

    # Totales generales
    ws[f'A{row}'] = "TOTAL INGRESOS:"
    total_ingresos = data.get('total_ingresos', 0)
    ws[f'B{row}'] = float(total_ingresos) if total_ingresos is not None else 0
    ws[f'B{row}'].number_format = '"S/ "#,##0.00'
    ws[f'A{row}'].font = Font(bold=True)

    row += 1
    ws[f'A{row}'] = "TOTAL EXÁMENES:"
    ws[f'B{row}'] = data.get('total_examenes', 0) or 0

    row += 1
    ws[f'A{row}'] = "INTERNAS:"
    total_internas = data.get('total_internas', 0)
    ws[f'B{row}'] = float(total_internas) if total_internas is not None else 0
    ws[f'B{row}'].number_format = '"S/ "#,##0.00'

    row += 1
    ws[f'A{row}'] = "AMBULATORIAS:"
    total_ambulatorias = data.get('total_ambulatorias', 0)
    ws[f'B{row}'] = float(total_ambulatorias) if total_ambulatorias is not None else 0
    ws[f'B{row}'].number_format = '"S/ "#,##0.00'

    # Ingresos por tipo de examen
    row += 2
    ws[f'A{row}'] = "INGRESOS POR TIPO DE EXAMEN"
    ws[f'A{row}'].font = Font(bold=True, italic=True)

    row += 1
    headers = ['Examen', 'Cantidad', 'Internas', 'Ambulatorias', 'Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    for item in data.get('examenes_por_tipo', []):
        row += 1

        # Examen
        ws.cell(row=row, column=1).value = item.get('examen_especifico', '')

        # Cantidad
        ws.cell(row=row, column=2).value = item.get('cantidad', 0) or 0

        # Internas
        val_internas = item.get('total_internas', 0)
        ws.cell(row=row, column=3).value = float(val_internas) if val_internas is not None else 0
        ws.cell(row=row, column=3).number_format = '"S/ "#,##0.00'

        # Ambulatorias
        val_ambulatorias = item.get('total_ambulatorias', 0)
        ws.cell(row=row, column=4).value = float(val_ambulatorias) if val_ambulatorias is not None else 0
        ws.cell(row=row, column=4).number_format = '"S/ "#,##0.00'

        # Total
        val_total = item.get('total_general', 0)
        ws.cell(row=row, column=5).value = float(val_total) if val_total is not None else 0
        ws.cell(row=row, column=5).number_format = '"S/ "#,##0.00'

        # Bordes
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border

    # Listado detallado
    row += 2
    ws[f'A{row}'] = "LISTADO DETALLADO DE EXÁMENES"
    ws[f'A{row}'].font = Font(bold=True, italic=True)

    row += 1
    headers = ['Fecha', 'Paciente', 'DNI', 'Examen', 'Tipo', 'Pago', 'Monto']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    for item in data.get('detalle', []):
        row += 1

        # Fecha
        ws.cell(row=row, column=1).value = f"{item.get('fecha', '')} {item.get('hora', '')}"

        # Paciente
        ws.cell(row=row, column=2).value = item.get('paciente', '')

        # DNI
        ws.cell(row=row, column=3).value = item.get('dni', '')

        # Examen
        ws.cell(row=row, column=4).value = item.get('examen', '')

        # Tipo
        ws.cell(row=row, column=5).value = item.get('tipo', '')

        # Pago
        ws.cell(row=row, column=6).value = item.get('metodo_pago', '')

        # Monto
        monto = item.get('monto', 0)
        ws.cell(row=row, column=7).value = float(monto) if monto is not None else 0
        ws.cell(row=row, column=7).number_format = '"S/ "#,##0.00'

        # Bordes
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border

    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
################
def exportar_farmacia(ws, data, header_font, header_fill, border):
    """Exportar farmacia a Excel"""
    row = 3
    
    # Totales
    ws[f'A{row}'] = "TOTAL INGRESOS:"
    total_ingresos = data.get('total_ingresos', 0)
    ws[f'B{row}'] = float(total_ingresos) if total_ingresos is not None else 0
    ws[f'B{row}'].number_format = '"S/ "#,##0.00'
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "TOTAL VENTAS:"
    ws[f'B{row}'] = data.get('total_ventas', 0)
    
    # Ventas por vendedor
    row += 2
    ws[f'A{row}'] = "VENTAS POR VENDEDOR"
    ws[f'A{row}'].font = Font(bold=True, italic=True)
    
    row += 1
    headers = ['Vendedor', 'Cantidad', 'Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    for v in data.get('ventas_por_vendedor', []):
        row += 1
        nombre = f"{v.get('usuario__first_name', '')} {v.get('usuario__last_name', '')}".strip()
        ws.cell(row=row, column=1).value = nombre or v.get('usuario__username', '')
        ws.cell(row=row, column=2).value = v.get('cantidad', 0)
        
        total_vendedor = v.get('total', 0)
        ws.cell(row=row, column=3).value = float(total_vendedor) if total_vendedor is not None else 0
        ws.cell(row=row, column=3).number_format = '"S/ "#,##0.00'
        
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = border
    
    # Top medicamentos
    row += 2
    ws[f'A{row}'] = "TOP 20 MEDICAMENTOS MÁS VENDIDOS"
    ws[f'A{row}'].font = Font(bold=True, italic=True)
    
    row += 1
    headers = ['Medicamento', 'Cantidad', 'Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    for item in data.get('ventas_por_medicamento', []):
        row += 1
        ws.cell(row=row, column=1).value = item.get('medicamento__nombre_comercial', '')
        ws.cell(row=row, column=2).value = item.get('cantidad', 0)
        
        total_med = item.get('total', 0)
        ws.cell(row=row, column=3).value = float(total_med) if total_med is not None else 0
        ws.cell(row=row, column=3).number_format = '"S/ "#,##0.00'
        
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = border
    
    # Listado detallado
    row += 2
    ws[f'A{row}'] = "LISTADO DETALLADO DE VENTAS"
    ws[f'A{row}'].font = Font(bold=True, italic=True)
    
    row += 1
    headers = ['Fecha', 'Vendedor', 'Medicamentos', 'Pago', 'Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    for item in data.get('detalle', []):
        row += 1
        ws.cell(row=row, column=1).value = f"{item.get('fecha', '')} {item.get('hora', '')}"
        ws.cell(row=row, column=2).value = item.get('vendedor', '')
        ws.cell(row=row, column=3).value = item.get('medicamentos', '')
        ws.cell(row=row, column=4).value = item.get('metodo_pago', '')
        
        total_item = item.get('total', 0)
        ws.cell(row=row, column=5).value = float(total_item) if total_item is not None else 0
        ws.cell(row=row, column=5).number_format = '"S/ "#,##0.00'
        
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border
    
    # Ajustar ancho
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
#################
def exportar_topico(ws, data, header_font, header_fill, border):
    """Exportar tópico a Excel"""
    row = 3
    
    ws[f'A{row}'] = "TOTAL INGRESOS:"
    total_ingresos = data.get('total_ingresos', 0)
    ws[f'B{row}'] = float(total_ingresos) if total_ingresos is not None else 0
    ws[f'B{row}'].number_format = '"S/ "#,##0.00'
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "TOTAL PROCEDIMIENTOS:"
    ws[f'B{row}'] = data.get('total_procedimientos', 0)
    
    # Procedimientos por tipo
    row += 2
    ws[f'A{row}'] = "PROCEDIMIENTOS MÁS REALIZADOS"
    ws[f'A{row}'].font = Font(bold=True, italic=True)
    
    row += 1
    headers = ['Procedimiento', 'Cantidad', 'Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    for item in data.get('procedimientos_por_tipo', []):
        row += 1
        ws.cell(row=row, column=1).value = item.get('descripcion', 'Procedimiento')
        ws.cell(row=row, column=2).value = item.get('cantidad', 0)
        
        total_proc = item.get('total', 0)
        ws.cell(row=row, column=3).value = float(total_proc) if total_proc is not None else 0
        ws.cell(row=row, column=3).number_format = '"S/ "#,##0.00'
        
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = border
    
    # Ventas por vendedor
    row += 2
    ws[f'A{row}'] = "VENTAS POR VENDEDOR"
    ws[f'A{row}'].font = Font(bold=True, italic=True)
    
    row += 1
    headers = ['Vendedor', 'Cantidad', 'Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    for v in data.get('ventas_por_vendedor', []):
        row += 1
        nombre = f"{v.get('usuario__first_name', '')} {v.get('usuario__last_name', '')}".strip()
        ws.cell(row=row, column=1).value = nombre or v.get('usuario__username', '')
        ws.cell(row=row, column=2).value = v.get('cantidad', 0)
        
        total_vendedor = v.get('total', 0)
        ws.cell(row=row, column=3).value = float(total_vendedor) if total_vendedor is not None else 0
        ws.cell(row=row, column=3).number_format = '"S/ "#,##0.00'
        
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = border
    
    # Listado detallado
    row += 2
    ws[f'A{row}'] = "LISTADO DETALLADO DE PROCEDIMIENTOS"
    ws[f'A{row}'].font = Font(bold=True, italic=True)
    
    row += 1
    headers = ['Fecha', 'Vendedor', 'Procedimientos', 'Pago', 'Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    for item in data.get('detalle', []):
        row += 1
        ws.cell(row=row, column=1).value = f"{item.get('fecha', '')} {item.get('hora', '')}"
        ws.cell(row=row, column=2).value = item.get('vendedor', '')
        ws.cell(row=row, column=3).value = item.get('procedimientos', '')
        ws.cell(row=row, column=4).value = item.get('metodo_pago', '')
        
        total_item = item.get('total', 0)
        ws.cell(row=row, column=5).value = float(total_item) if total_item is not None else 0
        ws.cell(row=row, column=5).number_format = '"S/ "#,##0.00'
        
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border
    
    # Ajustar ancho
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
################
def exportar_ingresos_totales(ws, data, header_font, header_fill, border):
    """Exportar ingresos totales a Excel"""
    row = 3
    
    headers = ['Módulo', 'Total', 'Cantidad']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    ws.cell(row=row, column=1).value = "Consultas"
    total_cons = data.get('total_consultas', 0)
    ws.cell(row=row, column=2).value = float(total_cons) if total_cons is not None else 0
    ws.cell(row=row, column=2).number_format = '"S/ "#,##0.00'
    ws.cell(row=row, column=3).value = data.get('cantidad_consultas', 0)
    
    row += 1
    ws.cell(row=row, column=1).value = "Laboratorio"
    total_lab = data.get('total_laboratorio', 0)
    ws.cell(row=row, column=2).value = float(total_lab) if total_lab is not None else 0
    ws.cell(row=row, column=2).number_format = '"S/ "#,##0.00'
    ws.cell(row=row, column=3).value = data.get('cantidad_lab', 0)
    
    row += 1
    ws.cell(row=row, column=1).value = "Farmacia"
    total_far = data.get('total_farmacia', 0)
    ws.cell(row=row, column=2).value = float(total_far) if total_far is not None else 0
    ws.cell(row=row, column=2).number_format = '"S/ "#,##0.00'
    ws.cell(row=row, column=3).value = data.get('cantidad_farmacia', 0)
    
    row += 1
    ws.cell(row=row, column=1).value = "Tópico"
    total_top = data.get('total_topico', 0)
    ws.cell(row=row, column=2).value = float(total_top) if total_top is not None else 0
    ws.cell(row=row, column=2).number_format = '"S/ "#,##0.00'
    ws.cell(row=row, column=3).value = data.get('cantidad_topico', 0)
    
    row += 2
    ws[f'A{row}'] = "TOTAL GENERAL"
    ws[f'A{row}'].font = Font(bold=True)
    total_gen = data.get('total_general', 0)
    ws[f'B{row}'] = float(total_gen) if total_gen is not None else 0
    ws[f'B{row}'].number_format = '"S/ "#,##0.00'
    ws[f'B{row}'].font = Font(bold=True)
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
################
def exportar_metodos_pago(ws, data, header_font, header_fill, border):
    """Exportar métodos de pago a Excel"""
    row = 3
    
    headers = ['Método', 'Consultas', 'Laboratorio', 'Farmacia', 'Tópico', 'Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    for item in data.get('metodos', []):
        row += 1
        ws.cell(row=row, column=1).value = item.get('metodo', '')
        
        # Consultas
        val_cons = item.get('consultas', 0)
        ws.cell(row=row, column=2).value = float(val_cons) if val_cons is not None else 0
        ws.cell(row=row, column=2).number_format = '"S/ "#,##0.00'
        
        # Laboratorio
        val_lab = item.get('laboratorio', 0)
        ws.cell(row=row, column=3).value = float(val_lab) if val_lab is not None else 0
        ws.cell(row=row, column=3).number_format = '"S/ "#,##0.00'
        
        # Farmacia
        val_far = item.get('farmacia', 0)
        ws.cell(row=row, column=4).value = float(val_far) if val_far is not None else 0
        ws.cell(row=row, column=4).number_format = '"S/ "#,##0.00'
        
        # Tópico
        val_top = item.get('topico', 0)
        ws.cell(row=row, column=5).value = float(val_top) if val_top is not None else 0
        ws.cell(row=row, column=5).number_format = '"S/ "#,##0.00'
        
        # Total
        val_total = item.get('total', 0)
        ws.cell(row=row, column=6).value = float(val_total) if val_total is not None else 0
        ws.cell(row=row, column=6).number_format = '"S/ "#,##0.00'
        
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border
    
    row += 2
    ws[f'A{row}'] = "TOTAL GENERAL"
    ws[f'A{row}'].font = Font(bold=True)
    total_gen = data.get('total_general', 0)
    ws[f'F{row}'] = float(total_gen) if total_gen is not None else 0
    ws[f'F{row}'].number_format = '"S/ "#,##0.00'
    ws[f'F{row}'].font = Font(bold=True)
    
    for col in range(1, 7):
        ws.column_dimensions[chr(64+col)].width = 15

################
def exportar_inventario(ws, data, header_font, header_fill, border):
    """Exportar inventario a Excel"""
    row = 3
    
    ws[f'A{row}'] = "RESUMEN GENERAL"
    ws[f'A{row}'].font = Font(bold=True, italic=True)
    
    row += 1
    ws[f'A{row}'] = "Total Productos:"
    ws[f'B{row}'] = data.get('total_productos', 0)
    
    row += 1
    ws[f'A{row}'] = "Total Unidades:"
    ws[f'B{row}'] = data.get('total_unidades', 0)
    
    row += 1
    ws[f'A{row}'] = "Valor Inventario:"
    valor_total = data.get('valor_total', 0)
    ws[f'B{row}'] = float(valor_total) if valor_total is not None else 0
    ws[f'B{row}'].number_format = '"S/ "#,##0.00'
    
    # Stock bajo
    row += 2
    ws[f'A{row}'] = f"STOCK BAJO ({data.get('total_stock_bajo', 0)})"
    ws[f'A{row}'].font = Font(bold=True, italic=True)
    
    row += 1
    headers = ['Código', 'Medicamento', 'Stock', 'Mínimo', 'Estado']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    for item in data.get('stock_bajo', []):
        row += 1
        ws.cell(row=row, column=1).value = item.get('codigo', '')
        ws.cell(row=row, column=2).value = item.get('nombre', '')
        ws.cell(row=row, column=3).value = item.get('stock_actual', 0)
        ws.cell(row=row, column=4).value = item.get('stock_minimo', 0)
        ws.cell(row=row, column=5).value = item.get('estado', '')
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border
    
    # Próximos a vencer
    row += 2
    ws[f'A{row}'] = f"PRÓXIMOS A VENCER ({data.get('total_proximos_vencer', 0)})"
    ws[f'A{row}'].font = Font(bold=True, italic=True)
    
    row += 1
    headers = ['Código', 'Medicamento', 'Vencimiento', 'Días', 'Estado']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    for item in data.get('proximos_vencer', []):
        row += 1
        ws.cell(row=row, column=1).value = item.get('codigo', '')
        ws.cell(row=row, column=2).value = item.get('nombre', '')
        ws.cell(row=row, column=3).value = item.get('fecha_vencimiento', '')
        ws.cell(row=row, column=4).value = item.get('dias', 0)
        ws.cell(row=row, column=5).value = item.get('estado', '')
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border
    
    # Ajustar ancho
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

#############
def get_compras(desde, hasta, page=1, request=None):
    """Reporte de compras a proveedores con filtro por estado"""
    from farmacia.models import Compra
    from django.core.paginator import Paginator
    from django.db.models import Sum, Q
    
    # Obtener filtro de estado desde el request
    estado = request.GET.get('estado', 'todas') if request else 'todas'
    
    print(f"🔍 Filtrando compras desde {desde} hasta {hasta}, estado: {estado}")
    
    # Base query
    compras = Compra.objects.filter(
        fecha_factura__gte=desde,
        fecha_factura__lte=hasta
    ).select_related('proveedor', 'usuario').order_by('-fecha_factura')
    
    # APLICAR FILTRO POR ESTADO
    if estado == 'pagadas':
        compras = compras.filter(observaciones__icontains='[FACTURA PAGADA]')
        print(f"✅ Filtro aplicado: pagadas")
    elif estado == 'pendientes':
        compras = compras.filter(observaciones__icontains='[FACTURA PENDIENTE]')
        print(f"✅ Filtro aplicado: pendientes")
    else:
        print(f"✅ Mostrando todas las compras")
    
    print(f"📊 Compras encontradas: {compras.count()}")
    
    # Calcular totales
    totales = compras.aggregate(
        total_subtotal=Sum('subtotal'),
        total_igv=Sum('igv'),
        total_general=Sum('total')
    )
    
    # Paginación
    paginator = Paginator(compras, 20)
    page_obj = paginator.get_page(page)
    
    # Preparar datos
    compras_data = []
    for c in page_obj:
        # Determinar estado desde observaciones
        if '[FACTURA PAGADA]' in c.observaciones:
            estado_pago = 'Pagada'
            badge_class = 'success'
        elif '[FACTURA PENDIENTE]' in c.observaciones:
            estado_pago = 'Pendiente'
            badge_class = 'danger'
        else:
            estado_pago = 'No especificado'
            badge_class = 'secondary'
        
        compras_data.append({
            'id': c.id,
            'numero_factura': c.numero_factura,
            'fecha_factura': c.fecha_factura.strftime('%d/%m/%Y'),
            'proveedor': c.proveedor.nombre,
            'ruc': c.proveedor.ruc,
            'subtotal': float(c.subtotal),
            'igv': float(c.igv),
            'total': float(c.total),
            'estado_pago': estado_pago,
            'badge_class': badge_class,
        })
    
    return {
        'compras': compras_data,
        'totales': {
            'subtotal': float(totales['total_subtotal'] or 0),
            'igv': float(totales['total_igv'] or 0),
            'total': float(totales['total_general'] or 0),
        },
        'paginator': paginator,
        'page_obj': page_obj,
    }


########excel funcional########
@login_required
@grupo_requerido('administrador')
def exportar_reporte_excel(request):
    """Exporta el reporte actual a Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from datetime import datetime
    from django.http import HttpResponse
    from io import BytesIO
    from decimal import Decimal
    from farmacia.models import Medicamento
    import traceback  # Para logging
    
    tipo_reporte = request.GET.get('tipo', 'consultas_periodo')
    fecha_desde = request.GET.get('desde', date.today().strftime('%Y-%m-%d'))
    fecha_hasta = request.GET.get('hasta', date.today().strftime('%Y-%m-%d'))

    try:
        desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    except:
        desde = date.today()
        hasta = date.today()

    # Crear archivo Excel base
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Reporte {tipo_reporte}"

    # Estilos
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    alignment_center = Alignment(horizontal='center')
    alignment_right = Alignment(horizontal='right')
    
    # ============================================
    # TRY-EXCEPT GENERAL PARA EVITAR ERRORES 500
    # ============================================
    try:
        # ============================================
        # REPORTE: DIGEMID
        # ============================================
        if tipo_reporte == 'digemid':
            try:
                # SOLO medicamentos con registro sanitario VÁLIDO Y STOCK > 0
                medicamentos = Medicamento.objects.filter(
                    activo=True,
                    registro_sanitario__isnull=False,
                    stock_actual__gt=0
                ).exclude(
                    registro_sanitario__exact=''
                ).exclude(
                    registro_sanitario__icontains='nan'
                ).exclude(
                    registro_sanitario__icontains='null'
                ).order_by('codigo')
                
                # Aplicar filtro de fechas
                if desde:
                    medicamentos = medicamentos.filter(fecha_creacion__date__gte=desde)
                if hasta:
                    medicamentos = medicamentos.filter(fecha_creacion__date__lte=hasta)
                
                total_registros = medicamentos.count()
                
                # Título principal
                ws.merge_cells('A1:D1')
                cell = ws['A1']
                cell.value = f"REPORTE DIGEMID - PRECIOS DE MEDICAMENTOS"
                cell.font = Font(bold=True, size=16)
                cell.alignment = alignment_center
                
                # Información del período
                ws.merge_cells('A2:D2')
                cell = ws['A2']
                cell.value = f"Período: {fecha_desde} al {fecha_hasta}"
                cell.font = Font(bold=True, size=12)
                cell.alignment = alignment_center
                
                # Total de registros
                ws.merge_cells('A3:D3')
                cell = ws['A3']
                cell.value = f"Total de medicamentos con registro sanitario y stock: {total_registros}"
                cell.font = Font(bold=True, size=11)
                cell.alignment = alignment_center
                
                # Fecha de generación
                ws.merge_cells('A4:D4')
                cell = ws['A4']
                cell.value = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                cell.font = Font(italic=True, size=10)
                cell.alignment = alignment_center
                
                # Encabezados
                headers = ['CodEstab', 'CodProd', 'Precio1 (Caja)', 'Precio2 (Unitario)']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=6, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = alignment_center
                
                # Escribir datos
                row_num = 7
                for m in medicamentos:
                    ws.cell(row=row_num, column=1, value='121875').border = border
                    ws.cell(row=row_num, column=1).alignment = alignment_center
                    ws.cell(row=row_num, column=2, value=m.codigo).border = border
                    ws.cell(row=row_num, column=2).alignment = alignment_center
                    
                    precio_caja = float(m.precio_por_caja) if m.precio_por_caja else 0
                    cell = ws.cell(row=row_num, column=3, value=precio_caja)
                    cell.border = border
                    cell.number_format = '#,##0.00'
                    cell.alignment = alignment_right
                    
                    precio_unitario = float(m.precio_venta) if m.precio_venta else 0
                    cell = ws.cell(row=row_num, column=4, value=precio_unitario)
                    cell.border = border
                    cell.number_format = '#,##0.00'
                    cell.alignment = alignment_right
                    
                    row_num += 1
                
                # Totales
                row_num += 1
                ws.merge_cells(f'A{row_num}:C{row_num}')
                ws.cell(row=row_num, column=1, value="TOTAL MEDICAMENTOS:").font = Font(bold=True)
                ws.cell(row=row_num, column=1).border = border
                ws.cell(row=row_num, column=1).alignment = alignment_right
                ws.cell(row=row_num, column=4, value=total_registros).font = Font(bold=True)
                ws.cell(row=row_num, column=4).border = border
                ws.cell(row=row_num, column=4).alignment = alignment_center
                
            except Exception as e:
                # Si falla DIGEMID, mostrar mensaje amigable
                print(f"Error en reporte DIGEMID: {str(e)}")
                traceback.print_exc()
                
                # Limpiar la hoja y mostrar mensaje de error
                ws.delete_rows(1, ws.max_row)
                
                ws.merge_cells('A1:D1')
                cell = ws['A1']
                cell.value = f"REPORTE DIGEMID - NO DISPONIBLE"
                cell.font = Font(bold=True, size=16, color="FF0000")
                cell.alignment = alignment_center
                
                ws.merge_cells('A3:D3')
                cell = ws['A3']
                cell.value = f"⚠️ El reporte DIGEMID no está disponible temporalmente"
                cell.font = Font(bold=True, size=14, color="FF0000")
                cell.alignment = alignment_center
                
                ws.merge_cells('A5:D5')
                cell = ws['A5']
                cell.value = f"Motivo: {str(e)[:100]}..." if len(str(e)) > 100 else str(e)
                cell.font = Font(italic=True, size=10)
                cell.alignment = alignment_center
                
                ws.merge_cells('A7:D7')
                cell = ws['A7']
                cell.value = f"Por favor, intenta con otro reporte o contacta al administrador"
                cell.font = Font(size=11)
                cell.alignment = alignment_center

        # ============================================
        # OTROS REPORTES CON SU PROPIO TRY-EXCEPT
        # ============================================
        elif tipo_reporte == 'consultas_periodo':
            try:
                data = get_consultas_periodo(desde, hasta)
                exportar_consultas_periodo(ws, data, header_font, header_fill, border)
            except Exception as e:
                print(f"Error en reporte consultas_periodo: {str(e)}")
                traceback.print_exc()
                mostrar_mensaje_error(ws, "CONSULTAS POR PERÍODO", str(e))
                
        elif tipo_reporte == 'consultas_doctor':
            try:
                data = get_consultas_doctor(desde, hasta)
                exportar_consultas_doctor(ws, data, header_font, header_fill, border)
            except Exception as e:
                print(f"Error en reporte consultas_doctor: {str(e)}")
                traceback.print_exc()
                mostrar_mensaje_error(ws, "CONSULTAS POR DOCTOR", str(e))
                
        elif tipo_reporte == 'laboratorio':
            try:
                data = get_laboratorio(desde, hasta)
                exportar_laboratorio(ws, data, header_font, header_fill, border)
            except Exception as e:
                print(f"Error en reporte laboratorio: {str(e)}")
                traceback.print_exc()
                mostrar_mensaje_error(ws, "LABORATORIO", str(e))
                
        elif tipo_reporte == 'farmacia':
            try:
                data = get_farmacia(desde, hasta)
                exportar_farmacia(ws, data, header_font, header_fill, border)
            except Exception as e:
                print(f"Error en reporte farmacia: {str(e)}")
                traceback.print_exc()
                mostrar_mensaje_error(ws, "FARMACIA", str(e))
                
        elif tipo_reporte == 'topico':
            try:
                data = get_topico(desde, hasta)
                exportar_topico(ws, data, header_font, header_fill, border)
            except Exception as e:
                print(f"Error en reporte topico: {str(e)}")
                traceback.print_exc()
                mostrar_mensaje_error(ws, "TÓPICO", str(e))
                
        elif tipo_reporte == 'ingresos_totales':
            try:
                data = get_ingresos_totales(desde, hasta)
                exportar_ingresos_totales(ws, data, header_font, header_fill, border)
            except Exception as e:
                print(f"Error en reporte ingresos_totales: {str(e)}")
                traceback.print_exc()
                mostrar_mensaje_error(ws, "INGRESOS TOTALES", str(e))
                
        elif tipo_reporte == 'metodos_pago':
            try:
                data = get_metodos_pago(desde, hasta)
                exportar_metodos_pago(ws, data, header_font, header_fill, border)
            except Exception as e:
                print(f"Error en reporte metodos_pago: {str(e)}")
                traceback.print_exc()
                mostrar_mensaje_error(ws, "MÉTODOS DE PAGO", str(e))
                
        elif tipo_reporte == 'inventario':
            try:
                data = get_inventario(request)
                exportar_inventario(ws, data, header_font, header_fill, border)
            except Exception as e:
                print(f"Error en reporte inventario: {str(e)}")
                traceback.print_exc()
                mostrar_mensaje_error(ws, "INVENTARIO", str(e))
                
        elif tipo_reporte == 'compras':
            try:
                data = get_compras(desde, hasta)
                exportar_compras(ws, data, header_font, header_fill, border)
            except Exception as e:
                print(f"Error en reporte compras: {str(e)}")
                traceback.print_exc()
                mostrar_mensaje_error(ws, "COMPRAS", str(e))
        
        # ============================================
        # REPORTE NO IMPLEMENTADO
        # ============================================
        else:
            ws.merge_cells('A3:C3')
            cell = ws['A3']
            cell.value = f"⚠️ Exportación para el reporte '{tipo_reporte}' no disponible"
            cell.font = Font(bold=True, color="FF0000", size=12)
            cell.alignment = alignment_center
            
            ws.merge_cells('A5:C5')
            cell = ws['A5']
            cell.value = "Esta funcionalidad estará disponible próximamente."
            cell.font = Font(italic=True, size=11)
            cell.alignment = alignment_center
    
    except Exception as e:
        # Error MUY grave (ni siquiera se pudo determinar el tipo de reporte)
        print(f"Error CRÍTICO en exportar_reporte_excel: {str(e)}")
        traceback.print_exc()
        
        # Limpiar hoja
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).value = None
        
        ws.merge_cells('A1:C1')
        cell = ws['A1']
        cell.value = f"ERROR INESPERADO"
        cell.font = Font(bold=True, size=16, color="FF0000")
        cell.alignment = alignment_center
        
        ws.merge_cells('A3:C3')
        cell = ws['A3']
        cell.value = f"No se pudo generar el reporte solicitado"
        cell.font = Font(bold=True, size=14)
        cell.alignment = alignment_center
        
        ws.merge_cells('A5:C5')
        cell = ws['A5']
        cell.value = f"Por favor, contacta al administrador del sistema"
        cell.font = Font(size=11)
        cell.alignment = alignment_center

    # Ajustar ancho de columnas
    for col in range(1, ws.max_column + 1):
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = adjusted_width

    # Preparar respuesta HTTP
    filename = f"reporte_{tipo_reporte}_{fecha_desde}_al_{fecha_hasta}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response

# ============================================
# FUNCIÓN AUXILIAR PARA MOSTRAR ERRORES
# ============================================
def mostrar_mensaje_error(ws, nombre_reporte, mensaje_error):
    """Limpia la hoja y muestra un mensaje de error amigable"""
    # Limpiar la hoja
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None
    
    ws.merge_cells('A1:D1')
    cell = ws['A1']
    cell.value = f"REPORTE {nombre_reporte.upper()} - NO DISPONIBLE"
    cell.font = Font(bold=True, size=16, color="FF0000")
    cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:D3')
    cell = ws['A3']
    cell.value = f"⚠️ El reporte no se pudo generar temporalmente"
    cell.font = Font(bold=True, size=14, color="FF0000")
    cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A5:D5')
    cell = ws['A5']
    cell.value = f"Motivo: {mensaje_error[:100]}..." if len(mensaje_error) > 100 else mensaje_error
    cell.font = Font(italic=True, size=10)
    cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A7:D7')
    cell = ws['A7']
    cell.value = f"Por favor, intenta con otro reporte o contacta al administrador"
    cell.font = Font(size=11)
    cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A9:D9')
    cell = ws['A9']
    cell.value = f"Reportes disponibles: Consultas, Laboratorio, Farmacia, Ingresos, etc."
    cell.font = Font(italic=True, size=10)
    cell.alignment = Alignment(horizontal='center')



####################
def get_reporte_digemid(desde, hasta, page=1):
    """
    Reporte DIGEMID - SOLO 4 COLUMNAS
    - CodEstab: 121875 (fijo)
    - CodProd: código del medicamento
    - Precio1: precio por caja
    - Precio2: precio de venta unitario
    - Filtros: 
        * Registro sanitario válido (no nulo, no vacío, no 'nan', no 'null')
        * Stock actual > 0 ✅ NUEVO
    """
    from farmacia.models import Medicamento
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    print(f"🔍 Generando Reporte DIGEMID desde {desde} hasta {hasta}")
    
    # ============================================
    # BASE QUERY CON TODOS LOS FILTROS
    # ============================================
    medicamentos = Medicamento.objects.filter(
        activo=True,
        registro_sanitario__isnull=False,  # No nulo
        stock_actual__gt=0                  # ✅ SOLO CON STOCK
    ).exclude(
        registro_sanitario__exact=''        # No vacío
    ).exclude(
        registro_sanitario__icontains='nan' # No contenga 'nan'
    ).exclude(
        registro_sanitario__icontains='null' # No contenga 'null'
    ).order_by('codigo')
    
    # Filtrar por rango de fechas (usando fecha_creacion)
    if desde:
        medicamentos = medicamentos.filter(fecha_creacion__date__gte=desde)
    if hasta:
        medicamentos = medicamentos.filter(fecha_creacion__date__lte=hasta)
    
    total_encontrados = medicamentos.count()
    print(f"📊 Medicamentos encontrados: {total_encontrados}")
    
    # ============================================
    # PAGINACIÓN
    # ============================================
    paginator = Paginator(medicamentos, 50)
    page_obj = paginator.get_page(page)
    
    # ============================================
    # PREPARAR DATOS - SOLO 4 COLUMNAS
    # ============================================
    medicamentos_data = []
    for m in page_obj:
        medicamentos_data.append({
            'CodEstab': '121875',
            'CodProd': m.codigo,
            'Precio1': float(m.precio_por_caja) if m.precio_por_caja else 0,
            'Precio2': float(m.precio_venta) if m.precio_venta else 0,
        })
    
    return {
        'medicamentos': medicamentos_data,
        'total': total_encontrados,
        'paginator': paginator,
        'page_obj': page_obj,
        'desde': desde.strftime('%d/%m/%Y') if desde else '',
        'hasta': hasta.strftime('%d/%m/%Y') if hasta else '',
    }